#!/usr/bin/env python3
from flask import Flask, render_template, request, send_file, flash, redirect, url_for, jsonify, Response
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import os
import re
import threading
import time
import pandas as pd
import tempfile
from pathlib import Path
import subprocess
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
from datetime import datetime, timedelta
import sys
from dotenv import load_dotenv
from bs4 import BeautifulSoup
import versus

# Load environment variables from .env file
load_dotenv()

app = Flask(__name__)
# Use environment variable for production, fallback for development
app.secret_key = os.environ.get('SECRET_KEY', 'dev_key_change_in_production_' + str(os.urandom(24).hex()))

# Security: CORS Protection (only allow your domain in production)
CORS(app, resources={
    r"/api/*": {
        "origins": os.environ.get('ALLOWED_ORIGINS', '*').split(',')
    }
})

# Security: Rate Limiting (prevent abuse)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"],
    storage_uri="memory://"
)

# Spotify API setup (using environment variables for credentials)
# Set SPOTIPY_CLIENT_ID and SPOTIPY_CLIENT_SECRET in environment
try:
    sp = spotipy.Spotify(auth_manager=SpotifyClientCredentials())
    SPOTIFY_ENABLED = True
except Exception as e:
    print(f"⚠️  Spotify API not configured: {e}")
    SPOTIFY_ENABLED = False

# Rate limiting disabled
DOWNLOAD_LIMIT = None
download_tracker = {}

# Auto-update chart data on startup (temporarily disabled for faster startup)
# print("Checking for chart data updates...")
# try:
#     # Use smart updater: tries Kaggle first, falls back to Billboard direct
#     result = subprocess.run(['python3', 'smart_update.py'],
#                           capture_output=True, text=True, timeout=120)
#     if result.returncode == 0:
#         print("✓ Data check complete!")
# except Exception as e:
#     print(f"⚠️  Could not check for updates: {e}")

# Find US Music Charts 100 data file
DATA_DIR = Path('data')
DESKTOP_PATH = Path.home() / 'Desktop' / 'hot100.csv'

# Try current directory first, then data directory, then Desktop
def find_data_file():
    # Check current directory
    for name in ['hot-100-current.csv', 'hot100.csv', 'billboard_hot_100.csv']:
        filepath = Path(name)
        if filepath.exists():
            return filepath
    # Check data directory
    if DATA_DIR.exists():
        for name in ['hot-100-current.csv', 'hot100.csv', 'billboard_hot_100.csv']:
            filepath = DATA_DIR / name
            if filepath.exists():
                return filepath
    # Check desktop
    if DESKTOP_PATH.exists():
        return DESKTOP_PATH
    raise FileNotFoundError("Chart data not found! Run auto_update_data.py first.")

BILLBOARD_DATA_PATH = find_data_file()

# Load data once at startup (cached)
print(f"Loading U.S. Top Charts 100 data from {BILLBOARD_DATA_PATH.name}...")
BILLBOARD_DATA = pd.read_csv(BILLBOARD_DATA_PATH, low_memory=False, dtype={'Artist': 'category', 'Song': 'category'})
print(f"Loaded {len(BILLBOARD_DATA)} records!")

# Load US Album Charts 200 data
BILLBOARD_200_PATH = Path('billboard200.csv')
if not BILLBOARD_200_PATH.exists() and (DATA_DIR / 'billboard200.csv').exists():
    BILLBOARD_200_PATH = DATA_DIR / 'billboard200.csv'
if BILLBOARD_200_PATH.exists():
    print(f"Loading U.S. Album Charts 200 data from {BILLBOARD_200_PATH.name}...")
    BILLBOARD_200_DATA = pd.read_csv(BILLBOARD_200_PATH, low_memory=False, dtype={'Artist': 'category', 'Song': 'category'})
    print(f"Loaded {len(BILLBOARD_200_DATA)} U.S. Album Charts 200 records!")
else:
    print("⚠️  US Album Charts 200 data not found. Album chart will be unavailable.")
    BILLBOARD_200_DATA = None

# Precompute request-invariant lookups once at startup (data is loaded once and never
# mutated in-process, so these are constant for the process lifetime).
_hot100_dates_parsed = pd.to_datetime(BILLBOARD_DATA['Date'], errors='coerce')
# Sorted, stripped, unique modern (1990+) artists for the autocomplete endpoint
_modern_raw = BILLBOARD_DATA.loc[_hot100_dates_parsed >= '1990-01-01', 'Artist'].str.strip().unique()
# Exclude combined collab credits ("X Featuring Y", "X & Y Duet With Z") — the
# base artist search already matches those rows, so they are noise in autocomplete.
_collab_markers = (' featuring ', ' feat. ', ' feat ', ' with ', ' x ', ' & ', ' + ', ' duet ', ' / ')
MODERN_ARTISTS = sorted(a for a in _modern_raw if not any(m in a.lower() for m in _collab_markers))
# Available Hot 100 chart dates (newest first) as display strings
HOT100_AVAILABLE_DATES = [pd.Timestamp(d).strftime('%Y-%m-%d') for d in sorted(_hot100_dates_parsed[_hot100_dates_parsed.dt.year >= 1958].dropna().unique(), reverse=True)]

if BILLBOARD_200_DATA is not None:
    _albums_dates_parsed = pd.to_datetime(BILLBOARD_200_DATA['Date'], errors='coerce')
    ALBUMS200_AVAILABLE_DATES = [pd.Timestamp(d).strftime('%Y-%m-%d') for d in sorted(_albums_dates_parsed[_albums_dates_parsed.dt.year >= 1963].dropna().unique(), reverse=True)]
else:
    ALBUMS200_AVAILABLE_DATES = []

# Billboard Global 200 / Global Excl. US (charts exist since Sept 2020)
def _load_global_chart(filename):
    path = DATA_DIR / filename
    if not path.exists():
        print(f"⚠️  {filename} not found. That global chart will be unavailable.")
        return None, []
    df = pd.read_csv(path, low_memory=False, dtype={'Artist': 'category', 'Song': 'category'})
    parsed = pd.to_datetime(df['Date'], errors='coerce')
    dates = [pd.Timestamp(d).strftime('%Y-%m-%d') for d in sorted(parsed.dropna().unique(), reverse=True)]
    print(f"Loaded {len(df)} records from {filename}")
    return df, dates

GLOBAL_200_DATA, GLOBAL200_AVAILABLE_DATES = _load_global_chart('global200.csv')
GLOBAL_XUS_DATA, GLOBALEXUS_AVAILABLE_DATES = _load_global_chart('globalexus.csv')
RADIO_DATA, RADIO_AVAILABLE_DATES = _load_global_chart('radio.csv')
DIGITAL_DATA, DIGITAL_AVAILABLE_DATES = _load_global_chart('digital_songs.csv')
STREAMING_DATA, STREAMING_AVAILABLE_DATES = _load_global_chart('streaming_songs.csv')
ARTIST100_DATA, ARTIST100_AVAILABLE_DATES = _load_global_chart('artist100.csv')
POP_AIRPLAY_DATA, POP_AIRPLAY_AVAILABLE_DATES = _load_global_chart('pop_airplay.csv')

# Genre/format airplay charts. These are still backfilling on first run; a chart
# whose CSV is absent loads as None and is hidden from the nav rather than
# erroring, so the app is usable while data lands.
ADULT_CONTEMPORARY_DATA, ADULT_CONTEMPORARY_AVAILABLE_DATES = _load_global_chart('adult_contemporary.csv')
ADULT_POP_DATA, ADULT_POP_AVAILABLE_DATES = _load_global_chart('adult_pop_airplay.csv')
RHYTHMIC_DATA, RHYTHMIC_AVAILABLE_DATES = _load_global_chart('rhythmic_airplay.csv')
COUNTRY_AIRPLAY_DATA, COUNTRY_AIRPLAY_AVAILABLE_DATES = _load_global_chart('country_airplay.csv')
ALTERNATIVE_DATA, ALTERNATIVE_AVAILABLE_DATES = _load_global_chart('alternative_airplay.csv')
RNB_HIPHOP_DATA, RNB_HIPHOP_AVAILABLE_DATES = _load_global_chart('rnb_hiphop_airplay.csv')
DANCE_AIRPLAY_DATA, DANCE_AIRPLAY_AVAILABLE_DATES = _load_global_chart('dance_airplay.csv')
ADULT_RNB_DATA, ADULT_RNB_AVAILABLE_DATES = _load_global_chart('adult_rnb_airplay.csv')

# Discontinued charts. Billboard stopped publishing these, so the weekly
# scraper never touches them and their newest week is fixed: Heatseekers Songs
# ends 2014-11-29, Dance Singles Sales 2007-02-24.
HEATSEEKERS_DATA, HEATSEEKERS_AVAILABLE_DATES = _load_global_chart('heatseekers_songs.csv')
DANCE_SALES_DATA, DANCE_SALES_AVAILABLE_DATES = _load_global_chart('dance_singles_sales.csv')

# Bubbling Under IS current, so unlike the two above it joins the weekly
# scraper. Its archive starts at the chart's 1992 revival, not its 1959 debut.
BUBBLING_DATA, BUBBLING_AVAILABLE_DATES = _load_global_chart('bubbling_under.csv')

# Canadian Hot 100. Billboard's archive starts 2007-03-31, earlier than the
# June 2007 launch the chart is usually credited with; the weeks from 03-31 on
# each carry their own date and a distinct ranking, so they are real rather
# than the pre-launch clamp that find_chart_start.py exists to detect.
CANADIAN_DATA, CANADIAN_AVAILABLE_DATES = _load_global_chart('canadian_hot100.csv')

# Hot Dance/Electronic Songs. Archive starts 2014-03-22 and runs unbroken to
# the current week. Its depth went DOWN over its life — 50 rows at launch, 25
# today — so the registry's 25 is the modern value only, not a scrape floor.
DANCE_ELECTRONIC_DATA, DANCE_ELECTRONIC_AVAILABLE_DATES = _load_global_chart('dance_electronic_songs.csv')

# Billboard Japan Hot 100. Archive starts 2011-04-09 at 10 rows and reaches 25
# today, so its rows-per-week legitimately varies. Six weeks are absent, each a
# single week and four of them early January: Billboard served a ranking
# identical to the week before, so the backfill's clamp guard dropped them
# rather than store a frozen chart twice under two dates.
JAPAN_DATA, JAPAN_AVAILABLE_DATES = _load_global_chart('japan_hot100.csv')

# The Official U.K. Singles Chart. Archive starts 2011-01-29 at 10 rows and
# doubles to 20 partway through, so rows-per-week legitimately varies here too.
# One week is absent, 2011-12-31: Billboard served the week before's ranking
# unchanged over New Year, and the backfill's clamp guard dropped the repeat
# rather than file one frozen chart under two dates.
UK_SONGS_DATA, UK_SONGS_AVAILABLE_DATES = _load_global_chart('uk_songs.csv')

# Hot Rock & Alternative Songs — Billboard's own <h1>, not "Rock Songs". Runs
# 2009-06-20 to date, 50 rows every single week, no gaps.
ROCK_SONGS_DATA, ROCK_SONGS_AVAILABLE_DATES = _load_global_chart('rock_songs.csv')

# Hot Gospel Songs. Runs 2005-03-19 to date, 25 rows every week, no gaps.
GOSPEL_DATA, GOSPEL_AVAILABLE_DATES = _load_global_chart('gospel_songs.csv')

# Hot Christian Songs. Runs 2003-06-21 to date with no gaps, but it GREW: 30
# rows for its first 148 weeks, then 40 for 165, and 50 since — so the
# registry's 50 is the modern depth only, not a scrape threshold.
CHRISTIAN_DATA, CHRISTIAN_AVAILABLE_DATES = _load_global_chart('christian_songs.csv')

# Hot Country Songs and Hot R&B/Hip-Hop Songs — the CONSUMPTION charts, which
# are different data from the existing country_airplay and rnb_hiphop airplay
# charts. Both start 1958-10-20 and are MONDAY-dated until 1961-12-25, and both
# date the Bicentennial week Sunday 1976-07-04; see CHART_CALENDARS in
# scripts/backfill_chart.py. Depth runs 29 to 100 across their lives.
COUNTRY_SONGS_DATA, COUNTRY_SONGS_AVAILABLE_DATES = _load_global_chart('country_songs.csv')
RNB_SONGS_DATA, RNB_SONGS_AVAILABLE_DATES = _load_global_chart('rnb_hiphop_songs.csv')

# Top Album Sales. Runs 1991-05-25 to date. Its depth SHRANK — 100 rows from
# 1991 through at least 2015, 50 today — so the registry's 50 is modern only.
TOP_ALBUM_SALES_DATA, TOP_ALBUM_SALES_AVAILABLE_DATES = _load_global_chart('top_album_sales.csv')

# Hot Latin Songs. Runs 1986-09-06 to date. Its early depth is the shallowest
# on the site — Billboard published ONE row a week for much of 1986 — before
# settling at 40 through the 1990s and 50 since 2005. Eleven weeks are absent:
# ten year-end freezes where Billboard repeated the prior week's ranking, and
# 1998-10-10/1998-10-17, which the site never published (both clamp forward to
# 1998-10-24 at source). The registry's 50 is the modern depth only.
LATIN_SONGS_DATA, LATIN_SONGS_AVAILABLE_DATES = _load_global_chart('latin_songs.csv')

# Mainstream Rock Airplay. Runs 1981-03-21 to date at 40 rows, the same depth at
# every era. Twelve weeks repeat the ranking of the week before them: Billboard
# published each one under its own date — the pages are full size and carry a
# matching 'Week of' heading — and simply froze the chart, mostly over New Year.
# They are listed in scripts/known_clamped_weeks.json. The one real hole is
# 1984-12-29, a week the chart skipped entirely; asking for it returns a page
# headed 1985-01-05, so the 14-day step there is correct and not data loss.
MAINSTREAM_ROCK_DATA, MAINSTREAM_ROCK_AVAILABLE_DATES = _load_global_chart('mainstream_rock.csv')

# Adult Alternative Airplay — Billboard's slug is `triple-a`, and neither
# `adult-alternative-songs` nor `adult-alternative-airplay` reaches it. Runs
# 1996-01-20 to date with no gaps. Its depth GREW, 20 rows to 30 to 40, so the
# registry's 40 is the modern value only and not a scrape threshold.
ADULT_ALT_DATA, ADULT_ALT_AVAILABLE_DATES = _load_global_chart('adult_alternative.csv')

# Rock & Alternative Airplay. A THIRD rock chart, distinct from both
# `alternative` (Alternative Airplay) and `rock_songs` (the consumption chart
# Hot Rock & Alternative Songs). Runs 2009-06-20 to date at 50 rows.
ROCK_AIRPLAY_DATA, ROCK_AIRPLAY_AVAILABLE_DATES = _load_global_chart('rock_airplay.csv')

# Christian Airplay — the airplay chart, NOT `christian` (Hot Christian Songs,
# the consumption chart already on the site). Runs 2003-06-21 to date.
CHRISTIAN_AIRPLAY_DATA, CHRISTIAN_AIRPLAY_AVAILABLE_DATES = _load_global_chart('christian_airplay.csv')

# Gospel Airplay — likewise distinct from `gospel` (Hot Gospel Songs). Runs
# 2005-03-19 to date at 30 rows.
GOSPEL_AIRPLAY_DATA, GOSPEL_AIRPLAY_AVAILABLE_DATES = _load_global_chart('gospel_airplay.csv')

# Latin Pop Airplay. Depth GREW 1 -> 15 -> 20 -> 25, so the registry's 25 is the
# modern value only; the chart really does serve ONE row a week for its first
# four weeks, the same shape as Hot Latin Songs in 1986. Two weeks are absent —
# 1998-10-10 and 1998-10-17, which Billboard never published on any Latin chart,
# leaving one 21-day step. Four weeks repeat the ranking before them and all four
# serve their own 'Week of' heading, so they are frozen weeks, not clamps; they
# are in scripts/known_clamped_weeks.json.
LATIN_POP_AIRPLAY_DATA, LATIN_POP_AIRPLAY_AVAILABLE_DATES = _load_global_chart('latin_pop_airplay.csv')

# ── Chart registry ──────────────────────────────────────────────────────────
# Single source of truth for chart metadata. The nav is built by looping this,
# so adding a chart cannot leave the nav out of sync — commit cd05690 had to
# touch 11 template files to add one chart's nav link.
#
# `depth` is the chart's CURRENT size, for display and y-axis scaling only. It
# must NOT be used as a scrape completeness threshold: these charts changed
# depth over their lifetimes (Adult Contemporary ran 19-20 rows in 1961 against
# 30 today), so see the floor in fast_billboard_scraper.py instead.
CHARTS = {
    'top100':      dict(label='The Hot 100™',   group='Songs',  depth=100, kind='song'),
    'global200':   dict(label='Global 200',       group='Songs',  depth=200, kind='song'),
    'globalexus':  dict(label='Global Excl. US',  group='Songs',  depth=200, kind='song'),
    'radio':       dict(label='Radio Songs',      group='Songs',  depth=40,  kind='song'),
    'digital':     dict(label='Digital Song Sales', group='Songs', depth=25, kind='song'),
    'streaming':   dict(label='Streaming Songs',  group='Songs',  depth=50,  kind='song'),

    'pop_airplay':        dict(label='Pop Airplay',         group='Airplay', depth=40, kind='song'),
    'adult_pop':          dict(label='Adult Pop Airplay',   group='Airplay', depth=40, kind='song'),
    'adult_contemporary': dict(label='Adult Contemporary',  group='Airplay', depth=30, kind='song'),
    'rhythmic':           dict(label='Rhythmic Airplay',    group='Airplay', depth=40, kind='song'),
    'country_airplay':    dict(label='Country Airplay',     group='Airplay', depth=60, kind='song'),
    'alternative':        dict(label='Alternative Airplay', group='Airplay', depth=40, kind='song'),
    'rnb_hiphop':         dict(label='Mainstream R&B/Hip-Hop', group='Airplay', depth=40, kind='song'),
    'dance_airplay':      dict(label='Dance Airplay',        group='Airplay', depth=40, kind='song'),
    'adult_rnb':          dict(label='Adult R&B Airplay',    group='Airplay', depth=30, kind='song'),
    'mainstream_rock':    dict(label='Mainstream Rock Airplay', group='Airplay', depth=40, kind='song'),
    'adult_alternative':  dict(label='Adult Alternative Airplay', group='Airplay', depth=40, kind='song'),
    'rock_airplay':       dict(label='Rock & Alternative Airplay', group='Airplay', depth=50, kind='song'),
    'christian_airplay':  dict(label='Christian Airplay',   group='Airplay', depth=40, kind='song'),
    'gospel_airplay':     dict(label='Gospel Airplay',      group='Airplay', depth=30, kind='song'),
    'latin_pop_airplay':  dict(label='Latin Pop Airplay',   group='Airplay', depth=25, kind='song'),

    'heatseekers':   dict(label='Heatseekers Songs',   group='Songs', depth=25, kind='song'),
    'dance_sales':   dict(label='Dance Singles Sales', group='Songs', depth=25, kind='song'),
    'bubbling':      dict(label='Bubbling Under Hot 100', group='Songs', depth=25, kind='song'),
    'canadian_hot100': dict(label='Canadian Hot 100', group='Songs', depth=100, kind='song'),
    'dance_electronic': dict(label='Hot Dance/Electronic Songs', group='Songs', depth=25, kind='song'),
    'japan_hot100':     dict(label='Billboard Japan Hot 100', group='Songs', depth=25, kind='song'),
    'uk_songs':         dict(label='The Official U.K. Singles Chart', group='Songs', depth=20, kind='song'),
    'rock_songs':       dict(label='Hot Rock & Alternative Songs', group='Songs', depth=50, kind='song'),
    'gospel':           dict(label='Hot Gospel Songs', group='Songs', depth=25, kind='song'),
    'christian':        dict(label='Hot Christian Songs', group='Songs', depth=50, kind='song'),
    'country_songs':    dict(label='Hot Country Songs', group='Songs', depth=50, kind='song'),
    'rnb_songs':        dict(label='Hot R&B/Hip-Hop Songs', group='Songs', depth=50, kind='song'),
    'latin_songs':      dict(label='Hot Latin Songs', group='Songs', depth=50, kind='song'),

    'albums200':   dict(label='Billboard 200™', group='Albums & Artists', depth=200, kind='album'),
    'artist100':   dict(label='Artist 100',       group='Albums & Artists', depth=100, kind='artist'),
    'top_album_sales': dict(label='Top Album Sales', group='Albums & Artists', depth=50, kind='album'),
}

# Frame + available dates per chart key. Keyed to the module globals above so
# existing code paths that reference them directly keep working unchanged.
CHART_DATA = {
    'top100':      (BILLBOARD_DATA,             HOT100_AVAILABLE_DATES),
    'global200':   (GLOBAL_200_DATA,            GLOBAL200_AVAILABLE_DATES),
    'globalexus':  (GLOBAL_XUS_DATA,            GLOBALEXUS_AVAILABLE_DATES),
    'radio':       (RADIO_DATA,                 RADIO_AVAILABLE_DATES),
    'digital':     (DIGITAL_DATA,               DIGITAL_AVAILABLE_DATES),
    'streaming':   (STREAMING_DATA,             STREAMING_AVAILABLE_DATES),
    'pop_airplay': (POP_AIRPLAY_DATA,           POP_AIRPLAY_AVAILABLE_DATES),
    'adult_pop':   (ADULT_POP_DATA,             ADULT_POP_AVAILABLE_DATES),
    'adult_contemporary': (ADULT_CONTEMPORARY_DATA, ADULT_CONTEMPORARY_AVAILABLE_DATES),
    'rhythmic':    (RHYTHMIC_DATA,              RHYTHMIC_AVAILABLE_DATES),
    'country_airplay': (COUNTRY_AIRPLAY_DATA,   COUNTRY_AIRPLAY_AVAILABLE_DATES),
    'alternative': (ALTERNATIVE_DATA,           ALTERNATIVE_AVAILABLE_DATES),
    'rnb_hiphop':  (RNB_HIPHOP_DATA,            RNB_HIPHOP_AVAILABLE_DATES),
    'dance_airplay': (DANCE_AIRPLAY_DATA,       DANCE_AIRPLAY_AVAILABLE_DATES),
    'adult_rnb':   (ADULT_RNB_DATA,             ADULT_RNB_AVAILABLE_DATES),
    'mainstream_rock':   (MAINSTREAM_ROCK_DATA, MAINSTREAM_ROCK_AVAILABLE_DATES),
    'adult_alternative': (ADULT_ALT_DATA,       ADULT_ALT_AVAILABLE_DATES),
    'rock_airplay':      (ROCK_AIRPLAY_DATA,    ROCK_AIRPLAY_AVAILABLE_DATES),
    'christian_airplay': (CHRISTIAN_AIRPLAY_DATA, CHRISTIAN_AIRPLAY_AVAILABLE_DATES),
    'gospel_airplay':    (GOSPEL_AIRPLAY_DATA,  GOSPEL_AIRPLAY_AVAILABLE_DATES),
    'latin_pop_airplay': (LATIN_POP_AIRPLAY_DATA, LATIN_POP_AIRPLAY_AVAILABLE_DATES),
    'heatseekers': (HEATSEEKERS_DATA,           HEATSEEKERS_AVAILABLE_DATES),
    'dance_sales': (DANCE_SALES_DATA,           DANCE_SALES_AVAILABLE_DATES),
    'bubbling':    (BUBBLING_DATA,              BUBBLING_AVAILABLE_DATES),
    'canadian_hot100': (CANADIAN_DATA,          CANADIAN_AVAILABLE_DATES),
    'dance_electronic': (DANCE_ELECTRONIC_DATA, DANCE_ELECTRONIC_AVAILABLE_DATES),
    'japan_hot100':    (JAPAN_DATA,             JAPAN_AVAILABLE_DATES),
    'uk_songs':        (UK_SONGS_DATA,          UK_SONGS_AVAILABLE_DATES),
    'rock_songs':      (ROCK_SONGS_DATA,        ROCK_SONGS_AVAILABLE_DATES),
    'gospel':          (GOSPEL_DATA,            GOSPEL_AVAILABLE_DATES),
    'christian':       (CHRISTIAN_DATA,         CHRISTIAN_AVAILABLE_DATES),
    'country_songs':   (COUNTRY_SONGS_DATA,     COUNTRY_SONGS_AVAILABLE_DATES),
    'rnb_songs':       (RNB_SONGS_DATA,         RNB_SONGS_AVAILABLE_DATES),
    'latin_songs':     (LATIN_SONGS_DATA,       LATIN_SONGS_AVAILABLE_DATES),
    'top_album_sales': (TOP_ALBUM_SALES_DATA,   TOP_ALBUM_SALES_AVAILABLE_DATES),
    'albums200':   (BILLBOARD_200_DATA,         ALBUMS200_AVAILABLE_DATES),
    'artist100':   (ARTIST100_DATA,             ARTIST100_AVAILABLE_DATES),
}

# Dates parsed once at startup. Every request previously re-parsed the whole
# Date column (123k rows on Adult Contemporary); the versus feature would have
# multiplied that per artist compared. Index-aligned to each frame so callers
# can mask with it directly.
CHART_DT = {}
for _k, (_df, _d) in CHART_DATA.items():
    if _df is not None and len(_df):
        CHART_DT[_k] = pd.to_datetime(_df['Date'], errors='coerce')
del _k, _df, _d

# Autocomplete pool per chart. MODERN_ARTISTS is Hot 100 rows from 1990+, which
# would never surface a country act who did not cross over, and would hide the
# pre-1990 half of Adult Contemporary. Precomputed once, like MODERN_ARTISTS.
CHART_ARTISTS = {}
for _k, (_df, _d) in CHART_DATA.items():
    if _df is None or not len(_df):
        continue
    _names = _df['Artist'].dropna().astype(str).str.strip().unique()
    CHART_ARTISTS[_k] = sorted(
        a for a in _names
        if a and not any(m in a.lower() for m in _collab_markers)
    )
del _k, _df, _d, _names

# Every artist on any chart, for the report search box. MODERN_ARTISTS is Hot
# 100 1990+ (2,943 names) — against 15,152 here, it made 12,209 artists
# unreachable from a search box that can now report on all of them.
ALL_ARTISTS = sorted({a for names in CHART_ARTISTS.values() for a in names})


# Year-end charts.
# One combined CSV rather than the per-chart files the weekly charts use: the
# whole year-end dataset is smaller than one weekly chart and every row has the
# same shape, so 20 more globals would buy nothing.
#
# Only genuine years are in the file. Billboard answers any year, clamping a
# missing one forward to the next year it holds, and a chart with no year-end
# edition at all is answered with its current weekly chart. Neither states a
# year anywhere on the page. See scripts/yearend_guard.py and
# docs/HANDOFF-year-end.md.
def _load_yearend():
    path = DATA_DIR / 'yearend.csv'
    if not path.exists():
        print('yearend.csv not found. Year-end views will be unavailable.')
        return {}, {}
    df = pd.read_csv(path, low_memory=False)
    df['Year'] = pd.to_numeric(df['Year'], errors='coerce')
    df['Rank'] = pd.to_numeric(df['Rank'], errors='coerce')
    df = df.dropna(subset=['Year', 'Rank'])
    df['Year'] = df['Year'].astype(int)
    df['Rank'] = df['Rank'].astype(int)

    data, years = {}, {}
    for key, group in df.groupby('Chart'):
        if key not in CHARTS:
            continue
        group = group.sort_values(['Year', 'Rank'])
        data[key] = group
        years[key] = sorted(group['Year'].unique().tolist(), reverse=True)
    print(f'Loaded {len(df)} year-end records across {len(data)} charts')
    return data, years


YEAREND_DATA, YEAREND_YEARS = _load_yearend()


def available_charts():
    """Registry entries that actually have data loaded, grouped for the nav."""
    groups = {}
    for key, meta in CHARTS.items():
        df, _ = CHART_DATA.get(key, (None, None))
        if df is None or not len(df):
            continue
        groups.setdefault(meta['group'], []).append(dict(meta, key=key))
    return groups

def yearend_charts():
    """Charts with a genuine year-end edition, in registry order.

    Built from YEAREND_YEARS rather than CHARTS on purpose: rnb_hiphop,
    heatseekers and bubbling have no year-end edition at all, so listing every
    chart here would put three links in the nav that only redirect back with an
    error. See docs/HANDOFF-year-end.md.
    """
    return [dict(CHARTS[key], key=key) for key in CHARTS if key in YEAREND_YEARS]

@app.context_processor
def inject_nav():
    """Every template gets the nav data, so the nav lives in one partial."""
    return {
        'nav_groups': available_charts(),
        'nav_charts': CHARTS,
        'nav_yearend': yearend_charts(),
    }

def safe_int(val, default=None):
    if pd.isna(val):
        return default
    if isinstance(val, str) and (val.strip() == '-' or val.strip() == ''):
        return default
    try:
        result = int(float(val))
        return result if result != 0 else default
    except (ValueError, TypeError):
        return default

# Credit parsing lives in versus.py — the one module that can be unit-tested
# without loading 1.5M rows. Song keying and artist selection have to agree
# about where one artist's name ends, so each has a single definition and
# these are aliases, not copies.
primary_artist = versus.primary_artist
artist_match_mask = versus.artist_match_mask
_CREDIT_MARKER_RE = versus._CREDIT_MARKER_RE


def _primary_artist_col(credits):
    """primary_artist over a whole credit column.

    _song_chart_page used to inline the marker regex here as a vectorized
    str.replace, which is exactly the copy the note above says must not exist,
    and it drifted: the inline version never learned to split on commas, so
    Billboard's two credits for 'Free' ('Rumi, JINU, EJAE & Andrew Choi' and
    'Rumi & JINU: EJAE & Andrew Choi') keyed differently here than in the
    scalar lookups a few lines below, and the song showed a 4-week run instead
    of 20. Mapping per UNIQUE credit keeps the speed that motivated the inline
    version — 1.5M rows resolve to ~30k distinct credits — without a copy.
    """
    stripped = credits.str.strip()
    lookup = {c: primary_artist(c) for c in stripped.dropna().unique()}
    return stripped.map(lookup)

def check_download_limit(ip_address):
    """Rate limiting disabled - always allow downloads"""
    return True, 0  # Always allowed

def process_billboard_data(artist_name):
    """Process Billboard data and return Excel file path"""
    import openpyxl

    data = BILLBOARD_DATA.copy()
    required_cols = {'Date', 'Song', 'Artist', 'Rank'}
    if not required_cols.issubset(data.columns):
        return None, "Missing required columns: Date, Song, Artist, Rank"

    data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
    data = data.dropna(subset=['Date'])
    data['Song_Clean'] = data['Song'].str.strip().str.title()
    data['Artist_Clean'] = data['Artist'].str.strip()

    filtered_data = data[artist_match_mask(data['Artist_Clean'], artist_name)].copy()

    if filtered_data.empty:
        return None, f"No results found for artist: {artist_name}"

    pivot = filtered_data.pivot_table(
        index='Date', columns='Song_Clean', values='Rank', aggfunc='first'
    )

    all_weeks = pd.date_range(start=filtered_data['Date'].min(), end=filtered_data['Date'].max(), freq='7D')
    pivot = pivot.reindex(all_weeks)

    first_appearance = filtered_data.groupby('Song_Clean')['Date'].min()
    sorted_cols = [c for c in first_appearance.sort_values().index if c in pivot.columns]
    pivot = pivot[sorted_cols]

    # Get first valid image URL per song
    has_images = 'Image URL' in filtered_data.columns
    def get_image(song):
        if not has_images:
            return None
        urls = filtered_data[filtered_data['Song_Clean'] == song]['Image URL'].dropna()
        urls = urls[urls != '#']
        return urls.iloc[0] if not urls.empty else None

    safe_name = artist_name.title().replace(" ", "_")
    output_file = os.path.join(tempfile.gettempdir(), f'{safe_name}_Chart_History.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    ws.cell(row=1, column=1, value='Song')
    for ci, song in enumerate(sorted_cols, start=2):
        ws.cell(row=1, column=ci, value=song)

    ws.cell(row=2, column=1, value='Image')
    for ci, song in enumerate(sorted_cols, start=2):
        ws.cell(row=2, column=ci, value=get_image(song))

    for ri, (date, row) in enumerate(pivot.iterrows(), start=3):
        cell = ws.cell(row=ri, column=1, value=date.to_pydatetime() if hasattr(date, 'to_pydatetime') else date)
        cell.number_format = 'M/D/YYYY'
        for ci, song in enumerate(sorted_cols, start=2):
            val = row[song]
            if pd.notna(val):
                ws.cell(row=ri, column=ci, value=int(val))

    wb.save(output_file)
    return output_file, None

def process_album_data_excel(artist_name):
    """Process Billboard 200 album data and return Excel file path"""
    import openpyxl

    if BILLBOARD_200_DATA is None:
        return None, "Album data not available"

    data = BILLBOARD_200_DATA.copy()
    required_cols = {'Date', 'Song', 'Artist', 'Rank'}
    if not required_cols.issubset(data.columns):
        return None, "Missing required columns: Date, Song, Artist, Rank"

    data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
    data = data.dropna(subset=['Date'])
    data['Album_Clean'] = data['Song'].str.strip().str.title()
    data['Artist_Clean'] = data['Artist'].str.strip()

    filtered_data = data[artist_match_mask(data['Artist_Clean'], artist_name)].copy()

    if filtered_data.empty:
        return None, f"No album results found for artist: {artist_name}"

    pivot = filtered_data.pivot_table(
        index='Date', columns='Album_Clean', values='Rank', aggfunc='first'
    )

    all_weeks = pd.date_range(start=filtered_data['Date'].min(), end=filtered_data['Date'].max(), freq='7D')
    pivot = pivot.reindex(all_weeks)

    first_appearance = filtered_data.groupby('Album_Clean')['Date'].min()
    sorted_cols = [c for c in first_appearance.sort_values().index if c in pivot.columns]
    pivot = pivot[sorted_cols]

    has_images = 'Image URL' in filtered_data.columns
    def get_image(album):
        if not has_images:
            return None
        urls = filtered_data[filtered_data['Album_Clean'] == album]['Image URL'].dropna()
        urls = urls[urls != '#']
        return urls.iloc[0] if not urls.empty else None

    safe_name = artist_name.title().replace(" ", "_")
    output_file = os.path.join(tempfile.gettempdir(), f'{safe_name}_Albums_Chart_History.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    ws.cell(row=1, column=1, value='Song')
    for ci, album in enumerate(sorted_cols, start=2):
        ws.cell(row=1, column=ci, value=album)

    ws.cell(row=2, column=1, value='Image')
    for ci, album in enumerate(sorted_cols, start=2):
        ws.cell(row=2, column=ci, value=get_image(album))

    for ri, (date, row) in enumerate(pivot.iterrows(), start=3):
        cell = ws.cell(row=ri, column=1, value=date.to_pydatetime() if hasattr(date, 'to_pydatetime') else date)
        cell.number_format = 'M/D/YYYY'
        for ci, album in enumerate(sorted_cols, start=2):
            val = row[album]
            if pd.notna(val):
                ws.cell(row=ri, column=ci, value=int(val))

    wb.save(output_file)
    return output_file, None

@app.route('/')
def index():
    return redirect('/top100')

@app.route('/search')
def search():
    return render_template('search.html')

@app.route('/about')
def about():
    return render_template('about.html')

def artist_chart_summaries(artist_name):
    """One artist's scorecard on every chart they appear on.

    Built on the versus primitives rather than new stat logic: the report and
    the versus scorecard must never be able to disagree about the same peak.

    `total_weeks_charted` is the "did they chart here" test, not `entries` —
    entries is nulled on artist charts, so it cannot distinguish "no rows"
    from "not a meaningful count here".
    """
    charts = []
    hidden = 0
    for key, meta in CHARTS.items():
        rows = _versus_artist_rows(key, artist_name)
        stats = versus.compute_artist_stats(rows, kind=meta['kind'])
        if not stats['total_weeks_charted']:
            hidden += 1
            continue
        stats.pop('timeline', None)
        charts.append({
            'key': key,
            'label': meta['label'],
            'group': meta['group'],
            'depth': meta['depth'],
            'kind': meta['kind'],
            **stats,
        })
    if not charts:
        return None
    return {'charts': charts, 'hidden': hidden}

def artist_chart_detail(artist_name, chart_key):
    """Per-entry rank history for one artist on one chart.

    The frame comes from CHART_DATA and there is no date floor — the version
    this replaced hardcoded the Hot 100 and cut everything before 1990, which
    silently truncated 46% of that chart's history.
    """
    meta = CHARTS.get(chart_key)
    df, _dates = CHART_DATA.get(chart_key, (None, None))
    if meta is None or df is None or not len(df):
        return None

    rows = df[artist_match_mask(df['Artist'], artist_name)].copy()
    if rows.empty:
        return None
    rows['Date'] = CHART_DT[chart_key].loc[rows.index]
    rows['Rank'] = pd.to_numeric(rows['Rank'], errors='coerce')
    rows = rows.dropna(subset=['Date', 'Rank'])
    if rows.empty:
        return None

    # Casing drifts week to week in the scraped data ("The Kid LAROI" vs
    # "The Kid Laroi"), so group case-insensitively and display the spelling
    # that appears most often.
    rows['Title_Clean'] = rows['Song'].astype(str).str.strip()
    rows['Title_Lower'] = rows['Title_Clean'].str.lower()
    titles = (
        rows.groupby('Title_Lower')['Title_Clean']
        .agg(lambda x: x.mode().iloc[0] if not x.mode().empty else x.iloc[0])
    )
    rows['Title'] = rows['Title_Lower'].map(titles)

    ordered = rows.sort_values('Date')
    series = {
        title: [{'date': d.strftime('%Y-%m-%d'), 'rank': int(r)}
                for d, r in zip(grp['Date'], grp['Rank'])]
        for title, grp in ordered.groupby('Title')
    }

    agg = rows.groupby('Title').agg(
        first_date=('Date', 'min'),
        weeks=('Date', 'nunique'),
        peak=('Rank', 'min'),
    )
    items = [
        {
            'name': title,
            'peak': int(row['peak']),
            'weeks': int(row['weeks']),
            'first_date': row['first_date'].strftime('%b %Y'),
            'first_date_sort': row['first_date'].strftime('%Y-%m-%d'),
            'is_number_one': int(row['peak']) == 1,
        }
        for title, row in agg.iterrows()
    ]
    items.sort(key=lambda x: (-x['weeks'], x['peak']))

    return {
        'chart': {'key': chart_key, 'label': meta['label'],
                  'kind': meta['kind'], 'depth': meta['depth']},
        'series': series,
        'items': items,
    }

@app.route('/analyze', methods=['POST'])
def analyze():
    artist_name = request.form.get('artist_name', '').strip()
    if not artist_name:
        flash('Please enter an artist name', 'error')
        return redirect(url_for('search'))

    try:
        summary = artist_chart_summaries(artist_name)
        if summary is None:
            flash(f'No results found for artist: {artist_name}', 'error')
            return redirect(url_for('search'))

        # Default to where they charted most, not always the Hot 100 — a
        # country act should not open on a chart they barely touched. entries
        # is None on artist charts, hence the `or 0`.
        selected = max(summary['charts'], key=lambda c: (c['entries'] or 0,
                                                         c['total_weeks_charted']))
        detail = artist_chart_detail(artist_name, selected['key'])

        return render_template(
            'results.html',
            artist_name=artist_name.title(),
            coverage=summary['charts'],
            hidden_count=summary['hidden'],
            selected_key=selected['key'],
            detail=detail,
        )
    except Exception as e:
        flash(f'An error occurred: {str(e)}', 'error')
        return redirect(url_for('search'))

@app.route('/api/artists')
@limiter.exempt
def get_artists():
    """API endpoint for artist autocomplete.

    With ?chart=<key> the pool is that chart's own artists — required by the
    versus page, where the Hot 100 pool would miss format-chart acts entirely.
    Without it, behaviour is unchanged for search.html.
    """
    query = request.args.get('q', '').lower()
    pool = CHART_ARTISTS.get(request.args.get('chart', ''), ALL_ARTISTS)

    if query:
        artists = [a for a in pool if a.lower().startswith(query)]
    else:
        artists = pool

    # Pools are pre-sorted, so slicing preserves order.
    return {'artists': list(artists[:50])}

# ── Artist versus ───────────────────────────────────────────────────────────

def _versus_artist_rows(chart_key, artist_name):
    """That artist's rows on one chart, with Date already parsed.

    Uses artist_match_mask, not substring matching: 'Tyla' must match
    'Tyla Featuring Zara Larsson' but not 'Tyla Yaweh' (commit 127a996).
    """
    df, _dates = CHART_DATA.get(chart_key, (None, None))
    if df is None or not len(df):
        return pd.DataFrame(columns=['Date', 'Rank', 'Song', 'Artist'])
    mask = artist_match_mask(df['Artist'], artist_name)
    # Image URL rides along when the chart has it so the CSV export can fill
    # its image row; the scorecard ignores the column.
    cols = ['Rank', 'Song', 'Artist']
    if 'Image URL' in df.columns:
        cols.append('Image URL')
    rows = df.loc[mask, cols].copy()
    rows['Date'] = CHART_DT[chart_key].loc[rows.index]
    return rows


def _parse_artist_list(raw):
    """Pipe-separated artists, blanks dropped, original order and case kept.

    Deduplicated case-insensitively, keeping the first-seen spelling: a
    repeated name (whether an exact repeat or just a different casing of the
    same query) is never a legitimate distinct comparison, and without this
    each repeat re-runs a full artist_match_mask pass over the chart.
    """
    seen = set()
    out = []
    for a in (raw or '').split('|'):
        a = a.strip()
        if not a:
            continue
        key = a.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(a)
    return out


@app.route('/api/versus')
@limiter.exempt
def api_versus():
    chart_key = request.args.get('chart', 'top100')
    meta = CHARTS.get(chart_key)
    if meta is None:
        return {'error': 'Unknown chart'}, 400
    df, _dates = CHART_DATA.get(chart_key, (None, None))
    if df is None or not len(df):
        return {'error': 'Chart data is not available'}, 400

    results = []
    for name in _parse_artist_list(request.args.get('artists')):
        rows = _versus_artist_rows(chart_key, name)
        stats = versus.compute_artist_stats(rows, kind=meta['kind'])
        # A typo in a four-artist comparison must not blank the page, so an
        # unmatched artist comes back with null stats rather than an error.
        results.append({
            'name': name,
            'display_name': versus.display_name(rows, name),
            **stats,
        })

    return {
        'chart': {
            'key': chart_key,
            'label': meta['label'],
            'depth': meta['depth'],
            'kind': meta['kind'],
        },
        'artists': results,
    }


@app.route('/api/artist-chart')
@limiter.exempt
def api_artist_chart():
    """One artist's detail on one chart, for the report's chart picker.

    The report lazy-loads this instead of inlining every chart: all 16 charts'
    detail is 519 KB on a heavy artist, against ~3 KB for the coverage table.
    """
    artist_name = (request.args.get('artist') or '').strip()
    if not artist_name:
        return {'error': 'Missing artist'}, 400
    chart_key = request.args.get('chart', 'top100')
    if chart_key not in CHARTS:
        return {'error': 'Unknown chart'}, 400
    detail = artist_chart_detail(artist_name, chart_key)
    if detail is None:
        return {'error': 'No data for artist on this chart'}, 404
    return detail


# Scorecard rows in the same order and with the same labels as the versus
# page's table, so a download matches what was on screen when it was clicked.
_VERSUS_CSV_ROWS = [
    ('Entries',       'entries'),
    ('Number ones',   'number_ones'),
    ('Weeks at #1',   'weeks_at_1'),
    ('Top 10s',       'top_10s'),
    ('Top 40s',       'top_40s'),
    ('Best peak',     'best_peak'),
    ('Weeks charted', 'total_weeks_charted'),
    ('First entry',   'first_entry'),
    ('Last entry',    'last_entry'),
    ('Biggest hit',   'biggest_hit'),
]


@app.route('/download-versus-csv')
@limiter.exempt
def download_versus_csv():
    """The on-screen comparison as CSV: the scorecard, then one row per chart
    week with each artist's best rank that week (blank when not charting).

    Same inputs as /api/versus, so the link is just the current URL's query
    string — a download always matches the comparison that produced it.
    """
    import io, csv as csvmod
    chart_key = request.args.get('chart', 'top100')
    meta = CHARTS.get(chart_key)
    if meta is None:
        return jsonify({'error': 'Unknown chart'}), 400
    names = _parse_artist_list(request.args.get('artists'))
    if not names:
        return jsonify({'error': 'No artists to compare'}), 400

    entries = []
    for name in names:
        rows = _versus_artist_rows(chart_key, name)
        entries.append((versus.display_name(rows, name),
                        versus.compute_artist_stats(rows, kind=meta['kind']),
                        rows))
    labels = [d for d, _, _ in entries]

    buf = io.StringIO()
    w = csvmod.writer(buf)
    w.writerow([f"{meta['label']} — Versus"])
    w.writerow([])
    w.writerow(['Stat'] + labels)
    for label, key in _VERSUS_CSV_ROWS:
        # A stat nulled for this chart kind writes blank, not 0 — the same
        # distinction the scorecard draws with an em dash.
        w.writerow([label] + [
            ('' if s.get(key) is None else s[key]) for _, s, _ in entries
        ])

    # One column per song, not one per artist. The previous layout wrote each
    # artist's best rank that week, which meant everything else they had
    # charting that week simply was not in the file. Each artist gets their own
    # block of song columns and the blocks share the date rows, so this reads
    # like the per-artist export with one extra row saying whose song a column
    # is. A collaboration between two compared artists appears in both blocks,
    # which is what you want when the point is to compare them.
    blocks = []
    for label, _stats, rows in entries:
        if rows.empty:
            continue
        pairs, headers, pivot, images = _song_pivot(rows)
        if pairs:
            blocks.append((label, pairs, headers, pivot, images))

    if blocks:
        w.writerow([])
        w.writerow(['Weekly rank by song'])
        song_row, artist_row, image_row = ['Song'], ['Artist'], ['Image']
        for label, pairs, headers, _pivot, images in blocks:
            song_row += headers
            artist_row += [label] * len(headers)
            image_row += [images.get(p, '') for p in pairs]
        w.writerow(song_row)
        w.writerow(artist_row)
        w.writerow(image_row)

        # Union of every week any of them charted, so the blocks line up on
        # shared dates and gaps stay visible as blanks.
        all_dates = sorted({d for _l, _p, _h, pv, _i in blocks for d in pv.index})
        width = sum(len(p) for _l, p, _h, _pv, _i in blocks)
        for date in _weeks_with_gaps(all_dates):
            if date is None:
                w.writerow([_GAP_MARKER] + [''] * width)
                continue
            row = [f'{date.month}/{date.day}/{date.year}']
            for _label, pairs, _headers, pivot, _images in blocks:
                if date in pivot.index:
                    ranks = pivot.loc[date]
                    row += [(int(ranks[p]) if pd.notna(ranks.get(p)) else '')
                            for p in pairs]
                else:
                    row += [''] * len(pairs)
            w.writerow(row)

    stem = '_vs_'.join(
        re.sub(r'[^\w\s-]', '', d).strip().replace(' ', '_') for d in labels[:3]
    ) or 'versus'
    return Response(buf.getvalue(), mimetype='text/csv', headers={
        'Content-Disposition': f'attachment; filename="{stem}_{chart_key}.csv"'})


def chart_dropouts(chart_key, week=None):
    """Entries on a chart's previous published week that are gone from its latest.

    Two things here are easy to get wrong.

    Matching is on a casefolded, stripped key, not the raw strings. Scraped
    artist casing drifts between weeks ("The Kid LAROI" vs "The Kid Laroi") --
    the same drift that once broke the new-vs-re-entry badge -- and an exact
    match would report every drift as a dropout plus a debut.

    "Previous week" is the previous PUBLISHED week from the date index, never
    the latest date minus seven days. Billboard skips weeks: holiday
    non-publication across most charts, and 11 weeks missing outright on Adult
    R&B. A fixed seven-day step would land on a week that does not exist, find
    nothing there, and report the entire chart as having dropped out.

    Artist charts carry the artist name in the Song column with Artist set to
    "Unknown", so they key on Song alone.
    """
    df, _ = CHART_DATA.get(chart_key, (None, None))
    dt = CHART_DT.get(chart_key)
    if df is None or dt is None or not len(df):
        return None

    weeks = sorted(dt.dropna().unique())
    if len(weeks) < 2:
        return None

    # `week` selects which week to report on; its comparison partner is always
    # the week immediately before it in this index, never a date arithmetic
    # step. The first week has no predecessor and so can never be reported on.
    index = len(weeks) - 1
    if week:
        wanted = pd.to_datetime(week, errors='coerce')
        if not pd.isna(wanted) and wanted in weeks:
            index = weeks.index(wanted)
    if index == 0:
        index = 1
    current_week, previous_week = weeks[index], weeks[index - 1]
    is_artist_chart = CHARTS[chart_key]['kind'] == 'artist'

    def entry_key(song, artist):
        title = str(song).casefold().strip()
        return title if is_artist_chart else (title, str(artist).casefold().strip())

    now = df[dt == current_week]
    still_charting = {entry_key(s, a) for s, a in zip(now['Song'], now['Artist'])}

    before = df[dt == previous_week].sort_values('Rank')
    gone = [
        {'title': song, 'artist': None if is_artist_chart else artist,
         'rank': safe_int(rank), 'weeks': safe_int(weeks_on)}
        for song, artist, rank, weeks_on in zip(
            before['Song'], before['Artist'], before['Rank'], before['Weeks on Chart'])
        if entry_key(song, artist) not in still_charting
    ]

    return {
        'key': chart_key,
        'label': CHARTS[chart_key]['label'],
        'group': CHARTS[chart_key]['group'],
        'current_week': current_week.strftime('%Y-%m-%d'),
        'previous_week': previous_week.strftime('%Y-%m-%d'),
        'previous_size': len(before),
        'dropouts': gone,
        'weeks': [w.strftime('%Y-%m-%d') for w in reversed(weeks[1:])],
    }


@app.route('/dropouts')
@limiter.exempt
def dropouts_page():
    """What fell off a chart in a given week. All state is in the URL so a
    particular week's dropouts are shareable and back/forward works."""
    chart_key = request.args.get('chart', 'top100')
    if chart_key not in CHARTS or CHART_DATA.get(chart_key, (None, None))[0] is None:
        chart_key = 'top100'
    report = chart_dropouts(chart_key, request.args.get('week'))
    return render_template(
        'dropouts.html',
        report=report,
        chart_key=chart_key,
        charts=available_charts(),
    )


@app.route('/versus')
@limiter.exempt
def versus_page():
    """Artist comparison scoped to one chart. All state is in the URL so
    comparisons are shareable and browser back/forward works."""
    chart_key = request.args.get('chart', 'top100')
    if chart_key not in CHARTS:
        chart_key = 'top100'
    return render_template(
        'versus.html',
        chart_key=chart_key,
        chart=CHARTS[chart_key],
        charts=available_charts(),
        initial_artists=_parse_artist_list(request.args.get('artists')),
    )

# ── Kworb cache (artist listeners + per-artist song streams) ────────────────
# Artist data: name_lower -> {'listeners': int, 'spotify_id': str}
_KWORB_CACHE = {'data': {}, 'ordered': [], 'fetched_at': None}
_KWORB_TTL = timedelta(hours=6)
_KWORB_URL = 'https://kworb.net/spotify/listeners.html'

# Per-artist songs: spotify_id -> {'fetched_at': dt, 'last_updated': str, 'songs': {song_lower: streams}}
_KWORB_SONGS_CACHE = {}
_KWORB_SONGS_URL = 'https://kworb.net/spotify/artist/{spotify_id}_songs.html'

def _refresh_kworb():
    import requests as _req
    try:
        resp = _req.get(_KWORB_URL, timeout=15, headers={'User-Agent': 'Mozilla/5.0 BillboardAnalyzer/1.0'})
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, 'html.parser')
        lookup = {}
        ordered = []  # rank-ordered list for top-N queries
        # Columns: rank, artist (link), listeners, daily change, peak rank, peak listeners
        for row in soup.select('table.sortable tbody tr'):
            tds = row.find_all('td')
            if len(tds) < 6:
                continue
            artist_tag = tds[1].find('a')
            if not artist_tag:
                continue
            name = artist_tag.get_text(strip=True)
            href = artist_tag.get('href', '')
            m = re.search(r'artist/([A-Za-z0-9]+)_songs\.html', href)
            spotify_id = m.group(1) if m else None
            try:
                listeners = int(tds[2].get_text(strip=True).replace(',', ''))
            except ValueError:
                continue
            try:
                peak_rank = int(tds[4].get_text(strip=True).replace(',', ''))
            except ValueError:
                peak_rank = None
            try:
                peak_listeners = int(tds[5].get_text(strip=True).replace(',', ''))
            except ValueError:
                peak_listeners = None
            rec = {
                'name': name,
                'listeners': listeners,
                'spotify_id': spotify_id,
                'peak_rank': peak_rank,
                'peak_listeners': peak_listeners,
            }
            lookup[name.lower()] = rec
            ordered.append(rec)
        _KWORB_CACHE['data'] = lookup
        _KWORB_CACHE['ordered'] = ordered
        _KWORB_CACHE['fetched_at'] = datetime.utcnow()
        print(f"✓ Kworb: loaded {len(lookup)} artist listener counts")
    except Exception as e:
        print(f"✗ Kworb fetch failed: {e}")

def get_top_artists(n=10):
    if not _KWORB_CACHE['fetched_at'] or datetime.utcnow() - _KWORB_CACHE['fetched_at'] > _KWORB_TTL:
        _refresh_kworb()
    return list(_KWORB_CACHE.get('ordered', []))[:n]

def _kworb_artist_record(artist_name):
    if not _KWORB_CACHE['fetched_at'] or datetime.utcnow() - _KWORB_CACHE['fetched_at'] > _KWORB_TTL:
        _refresh_kworb()
    lookup = _KWORB_CACHE['data']
    if not lookup:
        return None
    key = artist_name.lower()
    if key in lookup:
        return lookup[key]
    # Strip collab credit down to the primary artist (handles featuring/&/x/+/duet,
    # so remix credits like "A & B" or "A Featuring B, C & D" resolve to A). No
    # prefix fallback: it returned similarly-named artists ("Tyla" -> "Tyla Yaweh").
    clean = primary_artist(artist_name)
    if clean != key and clean in lookup:
        return lookup[clean]
    return None

def get_kworb_listeners(artist_name):
    rec = _kworb_artist_record(artist_name)
    return rec['listeners'] if rec else None

def _fetch_artist_songs(spotify_id):
    import requests as _req
    try:
        url = _KWORB_SONGS_URL.format(spotify_id=spotify_id)
        resp = _req.get(url, timeout=15, headers={'User-Agent': 'Mozilla/5.0 BillboardAnalyzer/1.0'})
        resp.raise_for_status()
        m = re.search(r'Last updated:\s*([\d/]+)', resp.text)
        last_updated = m.group(1) if m else None
        soup = BeautifulSoup(resp.text, 'html.parser')
        songs = {}
        # The songs table is the second table (class "addpos sortable").
        table = soup.select_one('table.addpos.sortable')
        if not table:
            return None
        for row in table.select('tbody tr'):
            tds = row.find_all('td')
            if len(tds) < 2:
                continue
            link = tds[0].find('a')
            if not link:
                continue
            title = link.get_text(strip=True)
            try:
                streams = int(tds[1].get_text(strip=True).replace(',', ''))
            except ValueError:
                continue
            daily = None
            if len(tds) >= 3:
                try:
                    daily = int(tds[2].get_text(strip=True).replace(',', ''))
                except ValueError:
                    pass
            songs[title.lower()] = {'streams': streams, 'daily': daily}
        return {'songs': songs, 'last_updated': last_updated, 'fetched_at': datetime.utcnow()}
    except Exception as e:
        print(f"✗ Kworb songs fetch failed for {spotify_id}: {e}")
        return None

def get_song_streams(artist_name, song_name):
    rec = _kworb_artist_record(artist_name)
    if not rec or not rec.get('spotify_id'):
        return None
    spotify_id = rec['spotify_id']
    cached = _KWORB_SONGS_CACHE.get(spotify_id)
    if not cached or datetime.utcnow() - cached['fetched_at'] > _KWORB_TTL:
        cached = _fetch_artist_songs(spotify_id)
        if cached:
            _KWORB_SONGS_CACHE[spotify_id] = cached
    if not cached:
        return None
    songs = cached['songs']
    key = song_name.lower().strip()
    base_key = re.sub(r'\s*[\(\[].*?[\)\]]\s*', ' ', key).strip()
    # Sum every version of the song (original + remixes/feat edits appear as
    # separate kworb rows like "Title", "Title (Remix)", "Title (feat. X)")
    streams = daily = None
    for stitle, sentry in songs.items():
        stitle_base = re.sub(r'\s*[\(\[].*?[\)\]]\s*', ' ', stitle).strip()
        if stitle == key or stitle_base == base_key:
            if sentry.get('streams') is not None:
                streams = (streams or 0) + sentry['streams']
            if sentry.get('daily') is not None:
                daily = (daily or 0) + sentry['daily']
    if streams is None:
        return None
    return {'streams': streams, 'daily': daily, 'last_updated': cached.get('last_updated')}

@app.route('/api/artist-info/<path:artist_name>')
def get_artist_info(artist_name):
    """API endpoint for artist information from Spotify (image) + Wikipedia/Billboard overview"""

    # Get Billboard data for statistics
    modern_data = BILLBOARD_DATA[pd.to_datetime(BILLBOARD_DATA['Date'], errors='coerce') >= '1990-01-01']
    artist_data = modern_data[modern_data['Artist'].str.strip().str.lower() == artist_name.lower()].copy()

    if artist_data.empty:
        return jsonify({'error': 'Artist not found in Billboard data'}), 404

    artist_name_proper = artist_data['Artist'].iloc[0].strip()

    # Calculate Billboard statistics
    artist_data['Date'] = pd.to_datetime(artist_data['Date'], errors='coerce')
    total_songs = artist_data['Song'].nunique()
    total_weeks = len(artist_data)
    rank_numeric = pd.to_numeric(artist_data['Rank'], errors='coerce')
    rank_valid = rank_numeric.dropna()
    peak_position = int(rank_valid.min()) if not rank_valid.empty else 0
    first_chart = artist_data['Date'].min().strftime('%B %Y')
    latest_chart = artist_data['Date'].max().strftime('%B %Y')
    # Best rank per song in one grouped pass, reused for both #1 and top-10 counts
    peak_per_song = artist_data.assign(_R=rank_numeric).groupby('Song', observed=True)['_R'].min()
    number_ones = int((peak_per_song == 1).sum())
    top_10_hits = int((peak_per_song <= 10).sum())

    # Create comprehensive Billboard-based description
    description_parts = []

    if number_ones > 0:
        if number_ones == 1:
            description_parts.append(f"Billboard Hot 100 chart-topper with {number_ones} #1 hit")
        else:
            description_parts.append(f"Billboard Hot 100 chart-topper with {number_ones} #1 hits")
    elif top_10_hits >= 5:
        description_parts.append(f"Billboard Hot 100 artist with {top_10_hits} top 10 hits")
    else:
        description_parts.append("Billboard Hot 100 charting artist")

    description_parts.append(f"{total_songs} songs charted for {total_weeks} total weeks")

    if first_chart != latest_chart:
        description_parts.append(f"Charting from {first_chart} to {latest_chart}")
    else:
        description_parts.append(f"First charted in {first_chart}")

    description = " • ".join(description_parts)

    # Initialize defaults
    image_url = None
    spotify_url = None
    overview = description  # Use Billboard description as default

    # Try to fetch from Wikipedia first (image + description)
    try:
        import requests

        # Try multiple Wikipedia search variations for disambiguation
        wiki_attempts = [
            artist_name_proper,  # Original name
            f"{artist_name_proper} (musician)",
            f"{artist_name_proper} (rapper)",
            f"{artist_name_proper} (singer)",
            f"{artist_name_proper} (band)"
        ]

        for attempt_name in wiki_attempts:
            try:
                wiki_url = f"https://en.wikipedia.org/api/rest_v1/page/summary/{requests.utils.quote(attempt_name)}"
                response = requests.get(wiki_url, timeout=10, headers={'User-Agent': 'Mozilla/5.0 BillboardAnalyzer/1.0'})

                if response.status_code == 200:
                    data = response.json()

                    # Check if this is a disambiguation page
                    page_type = data.get('type', '')
                    if page_type == 'disambiguation':
                        print(f"✗ Disambiguation page for {attempt_name}, trying next...")
                        continue

                    print(f"✓ Wikipedia page found for {attempt_name}")

                    # Get Wikipedia image (try multiple sources)
                    if 'originalimage' in data and data['originalimage'] and 'source' in data['originalimage']:
                        image_url = data['originalimage']['source']
                        print(f"✓ Found originalimage: {image_url}")
                    elif 'thumbnail' in data and data['thumbnail'] and 'source' in data['thumbnail']:
                        # Use larger thumbnail - replace size in URL
                        thumb_url = data['thumbnail']['source']
                        image_url = thumb_url.replace('/50px-', '/600px-').replace('/100px-', '/600px-').replace('/200px-', '/600px-').replace('/300px-', '/600px-')
                        print(f"✓ Found thumbnail (enlarged): {image_url}")

                    # Get Wikipedia description
                    if 'extract' in data and data['extract']:
                        wiki_extract = data['extract']
                        # Use first 2 sentences from Wikipedia
                        sentences = wiki_extract.split('. ')
                        if len(sentences) >= 2:
                            overview = '. '.join(sentences[:2]) + '.'
                        else:
                            overview = wiki_extract
                        print(f"✓ Got Wikipedia description: {len(overview)} chars")

                    # Found valid page, stop trying
                    break

            except Exception as inner_e:
                print(f"✗ Error trying {attempt_name}: {inner_e}")
                continue

        print(f"Wikipedia final: image_url={image_url is not None}, overview_len={len(overview) if overview else 0}")
    except Exception as e:
        print(f"✗ Wikipedia API error for {artist_name_proper}: {e}")

    # Try to get Spotify data as supplement if available
    spotify_followers = None
    spotify_popularity = None
    spotify_genres = []
    if SPOTIFY_ENABLED:
        try:
            results = sp.search(q=artist_name, type='artist', limit=1)
            if results['artists']['items']:
                artist_obj = results['artists']['items'][0]
                spotify_url = (artist_obj.get('external_urls') or {}).get('spotify')
                followers_obj = artist_obj.get('followers') or {}
                spotify_followers = followers_obj.get('total')
                spotify_popularity = artist_obj.get('popularity')
                spotify_genres = (artist_obj.get('genres') or [])[:4]

                # Use Spotify image if Wikipedia didn't provide one
                if not image_url and artist_obj['images']:
                    image_url = artist_obj['images'][0]['url']
        except Exception as e:
            print(f"Spotify API error: {e}")

    # If Spotify isn't available, try to construct Spotify URL from iTunes Search API
    if not spotify_url:
        try:
            from urllib.parse import quote

            # Try iTunes Search API which sometimes has Spotify artist IDs or links
            itunes_url = f"https://itunes.apple.com/search?term={quote(artist_name_proper)}&entity=allArtist&limit=1"
            itunes_response = requests.get(itunes_url, timeout=10, headers={'User-Agent': 'BillboardAnalyzer/1.0'})

            if itunes_response.status_code == 200:
                itunes_data = itunes_response.json()
                if itunes_data.get('results') and len(itunes_data['results']) > 0:
                    artist_result = itunes_data['results'][0]

                    # Get artist name from iTunes for potential Spotify search
                    itunes_artist_name = artist_result.get('artistName', artist_name_proper)

                    # Try to construct Spotify search URL (opens Spotify with search)
                    # Format: https://open.spotify.com/search/{artist_name}
                    search_query = quote(itunes_artist_name)
                    spotify_url = f"https://open.spotify.com/search/{search_query}"
                    print(f"✓ Created Spotify search URL: {spotify_url}")
        except Exception as e:
            print(f"iTunes/Spotify URL construction error: {e}")

    kworb_rec = _kworb_artist_record(artist_name_proper) or {}
    monthly_listeners = kworb_rec.get('listeners')
    listeners_peak = kworb_rec.get('peak_listeners')
    listeners_peak_rank = kworb_rec.get('peak_rank')

    return jsonify({
        'name': artist_name_proper,
        'image_url': image_url,
        'spotify_url': spotify_url,
        'spotify_followers': spotify_followers,
        'spotify_popularity': spotify_popularity,
        'spotify_genres': spotify_genres,
        'monthly_listeners': monthly_listeners,
        'listeners_peak': listeners_peak,
        'listeners_peak_rank': listeners_peak_rank,
        'overview': overview,
        'stats': {
            'total_songs': total_songs,
            'total_weeks': total_weeks,
            'peak_position': peak_position,
            'number_ones': number_ones,
            'top_10_hits': top_10_hits
        }
    })

@app.route('/api/top-artists')
@limiter.exempt
def get_top_artists_api():
    """Top N artists by Spotify monthly listeners (from kworb)."""
    try:
        n = max(1, min(int(request.args.get('n', 10)), 50))
    except (TypeError, ValueError):
        n = 10
    artists = get_top_artists(n)
    return jsonify({'artists': [
        {'name': a['name'], 'listeners': a['listeners'], 'peak_rank': a.get('peak_rank')}
        for a in artists
    ]})

@app.route('/api/current-number-ones')
@limiter.exempt
def get_current_number_ones():
    """Current #1 song (Hot 100) and #1 album (Billboard 200)."""
    out = {'song': None, 'album': None}
    try:
        d = BILLBOARD_DATA.copy()
        d['Date'] = pd.to_datetime(d['Date'], errors='coerce')
        d = d.dropna(subset=['Date', 'Rank'])
        latest = d['Date'].max()
        top = d[(d['Date'] == latest) & (d['Rank'] == 1)]
        if not top.empty:
            r = top.iloc[0]
            out['song'] = {
                'title': str(r['Song']).strip(),
                'artist': str(r['Artist']).strip(),
                'date': latest.strftime('%Y-%m-%d'),
                'weeks': int(d[(d['Song'].str.strip() == str(r['Song']).strip()) &
                                (d['Artist'].str.strip() == str(r['Artist']).strip())].shape[0]),
            }
    except Exception as e:
        print(f"current #1 song error: {e}")
    try:
        if BILLBOARD_200_DATA is not None:
            d = BILLBOARD_200_DATA.copy()
            d['Date'] = pd.to_datetime(d['Date'], errors='coerce')
            d = d.dropna(subset=['Date', 'Rank'])
            latest = d['Date'].max()
            top = d[(d['Date'] == latest) & (d['Rank'] == 1)]
            if not top.empty:
                r = top.iloc[0]
                out['album'] = {
                    'title': str(r.get('Album', r.get('Song', ''))).strip(),
                    'artist': str(r['Artist']).strip(),
                    'date': latest.strftime('%Y-%m-%d'),
                }
    except Exception as e:
        print(f"current #1 album error: {e}")
    return jsonify(out)

@app.route('/api/song-streams/<path:artist_name>/<path:song_name>')
@limiter.exempt
def get_song_streams_api(artist_name, song_name):
    """Spotify total stream count for a song, scraped from kworb."""
    result = get_song_streams(artist_name, song_name)
    if not result:
        return jsonify({'streams': None, 'last_updated': None})
    return jsonify(result)

# Deezer artwork lookups: results never change for a given song/album, but every
# chart page fires ~100 of them at once and Deezer's public quota is ~50 requests
# per 5s per IP. Cache successes forever and retry once when the quota trips
# (Deezer signals it with HTTP 200 + {"error": {"code": 4}}, not a 429).
_deezer_cache = {}
_deezer_cache_lock = threading.Lock()
_DEEZER_CACHE_MAX = 20000

def _deezer_cache_get(key):
    with _deezer_cache_lock:
        return _deezer_cache.get(key)

def _deezer_cache_put(key, value):
    with _deezer_cache_lock:
        if len(_deezer_cache) >= _DEEZER_CACHE_MAX:
            _deezer_cache.pop(next(iter(_deezer_cache)))
        _deezer_cache[key] = value

def _deezer_search_items(url):
    """GET a Deezer search URL, retrying once if the rate-limit quota trips."""
    import requests
    for attempt in (1, 2):
        response = requests.get(url, timeout=5)
        if response.status_code != 200:
            return []
        data = response.json()
        if data.get('error'):
            if attempt == 1:
                time.sleep(2.5)
                continue
            return []
        return data.get('data', [])
    return []

def _deezer_fallback_query(title, artist_credit):
    """Bare title + primary artist, for when the full Billboard credit string
    ('Shakira X Burna Boy', '... (FIFA World Cup Official Song 2026)') has no
    Deezer match."""
    bare_title = re.sub(r'\s*\([^)]*\)', '', title).strip() or title
    return f"{bare_title} {primary_artist(artist_credit)}"

def _deezer_track_payload(artist_name, song_name):
    """Deezer artwork for one track as the API's payload dict, or None.

    Split out of the route so the CSV exports can reach the same lookup and,
    more importantly, the same _deezer_cache: nine of the charts store no
    Image URL column at all, so their exports resolve artwork here instead.
    """
    cache_key = ('track', artist_name.casefold(), song_name.casefold())
    cached = _deezer_cache_get(cache_key)
    if cached:
        return cached

    from urllib.parse import quote

    query = f"{song_name} {artist_name}"
    url = f"https://api.deezer.com/search/track?q={quote(query)}&limit=3"

    items = _deezer_search_items(url)
    if not items:
        simple = _deezer_fallback_query(song_name, artist_name)
        if simple.casefold() != query.casefold():
            items = _deezer_search_items(f"https://api.deezer.com/search/track?q={quote(simple)}&limit=3")
    for item in items:
        cover = item.get('album', {}).get('cover_xl') or item.get('album', {}).get('cover_big', '')
        if cover:
            payload = {
                'image_url': cover,
                'track_name': item.get('title', ''),
                'source': 'deezer'
            }
            _deezer_cache_put(cache_key, payload)
            return payload
    return None


@app.route('/api/song-image/<path:artist_name>/<path:song_name>')
@limiter.exempt
def get_song_image(artist_name, song_name):
    """API endpoint to get song artwork from Deezer API"""
    try:
        payload = _deezer_track_payload(artist_name, song_name)
        if payload:
            return jsonify(payload)
        return jsonify({'error': 'Track not found'}), 404

    except Exception as e:
        print(f"Deezer API error for '{song_name}' by {artist_name}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/song-preview/<path:artist_name>/<path:song_name>')
@limiter.exempt
def get_song_preview(artist_name, song_name):
    """API endpoint to get a 30-second audio preview from Deezer API"""
    try:
        import requests
        from urllib.parse import quote

        query = f"{song_name} {artist_name}"
        url = f"https://api.deezer.com/search/track?q={quote(query)}&limit=3"
        response = requests.get(url, timeout=5)

        if response.status_code == 200:
            items = response.json().get('data', [])
            for item in items:
                preview = item.get('preview')
                if preview:
                    return jsonify({
                        'preview_url': preview,
                        'track_name': item.get('title', ''),
                        'source': 'deezer'
                    })

        return jsonify({'error': 'Preview not found'}), 404

    except Exception as e:
        print(f"Deezer preview error for '{song_name}' by {artist_name}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/album-image/<path:artist_name>/<path:album_name>')
@limiter.exempt
def get_album_image(artist_name, album_name):
    """API endpoint to get album artwork from Deezer API"""

    cache_key = ('album', artist_name.casefold(), album_name.casefold())
    cached = _deezer_cache_get(cache_key)
    if cached:
        return jsonify(cached)

    try:
        from urllib.parse import quote

        query = f"{album_name} {artist_name}"
        url = f"https://api.deezer.com/search/album?q={quote(query)}&limit=3"

        items = _deezer_search_items(url)
        if not items:
            simple = _deezer_fallback_query(album_name, artist_name)
            if simple.casefold() != query.casefold():
                items = _deezer_search_items(f"https://api.deezer.com/search/album?q={quote(simple)}&limit=3")
        for item in items:
            cover = item.get('cover_xl') or item.get('cover_big', '')
            if cover:
                payload = {
                    'image_url': cover,
                    'album_name': item.get('title', ''),
                    'artist_name': item.get('artist', {}).get('name', ''),
                    'source': 'deezer'
                }
                _deezer_cache_put(cache_key, payload)
                return jsonify(payload)

        return jsonify({'error': 'Album not found'}), 404

    except Exception as e:
        print(f"Deezer API error for album '{album_name}' by {artist_name}: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/artist-image/<path:artist_name>')
@limiter.exempt
def get_artist_image(artist_name):
    """Artist photo from Deezer (for the Artist 100 chart, whose entries are artists)."""
    cache_key = ('artist', artist_name.casefold())
    cached = _deezer_cache_get(cache_key)
    if cached:
        return jsonify(cached)
    try:
        from urllib.parse import quote
        # Search the primary artist so credits like "Drake & 21 Savage" still resolve
        for query in (artist_name, primary_artist(artist_name)):
            items = _deezer_search_items(f"https://api.deezer.com/search/artist?q={quote(query)}&limit=1")
            if items:
                pic = items[0].get('picture_xl') or items[0].get('picture_big', '')
                if pic:
                    payload = {'image_url': pic, 'artist_name': items[0].get('name', ''), 'source': 'deezer'}
                    _deezer_cache_put(cache_key, payload)
                    return jsonify(payload)
        return jsonify({'error': 'Artist not found'}), 404
    except Exception as e:
        print(f"Deezer artist-image error for '{artist_name}': {e}")
        return jsonify({'error': str(e)}), 500

# Billboard lazy-loads chart artwork, so a row scraped before its <img> swapped
# in the real source carries this placeholder gif instead. It is not artwork,
# and 16% of the year-end scrape has it.
_BB_LAZY_FALLBACK = 'lazyload-fallback'


def _billboard_art(raw, size=''):
    """Billboard's own artwork URL for a year-end row, or '' if it has none.

    charts-static.billboard.com is what billboard.com itself serves chart
    thumbnails from, and it hotlinks without a referer. It publishes exactly
    two sizes — 180x180 (as scraped) and 344x344; every other dimension 403s,
    so `size` may only ever be one of those two.
    """
    if not raw or _BB_LAZY_FALLBACK in raw:
        return ''
    if size and '-180x180.' in raw:
        return raw.replace('-180x180.', f'-{size}.')
    return raw


def _yearend_chart_page(chart_key):
    """Year-end view of a chart.

    Year-end rows have no Last Week, Peak Position or Weeks on Chart, so none
    of the weekly badges, filters or history apply — the template hides them
    under mode='yearend'.

    A year not in YEAREND_YEARS is one Billboard fabricates, so it redirects
    rather than rendering: a shared link must not be able to produce a page of
    invented history.
    """
    years = YEAREND_YEARS.get(chart_key)
    df = YEAREND_DATA.get(chart_key)
    if not years or df is None:
        flash(f"{CHARTS[chart_key]['label']} has no year-end chart", 'error')
        return redirect(url_for(chart_key))

    raw = request.args.get('year')
    selected_year = years[0]
    if raw:
        try:
            asked = int(raw)
        except ValueError:
            flash('Invalid year', 'error')
            return redirect(url_for(chart_key, view='yearend'))
        if asked not in years:
            flash(f'Billboard publishes no {asked} year-end chart for '
                  f"{CHARTS[chart_key]['label']}", 'error')
            return redirect(url_for(chart_key, view='yearend'))
        selected_year = asked

    rows = df[df['Year'] == selected_year].sort_values('Rank')
    is_artist_chart = CHARTS.get(chart_key, {}).get('kind') == 'artist'

    def _row(r):
        name = str(r['Song']).strip() if pd.notna(r['Song']) else ''
        credit = str(r['Artist']).strip() if pd.notna(r['Artist']) else ''
        art = str(r['Image URL']) if pd.notna(r['Image URL']) else ''
        return {
            'rank': int(r['Rank']),
            'song': name,
            # Artist 100 year-end rows are artists, not songs: the name sits in
            # the Song column and Artist is the literal string 'Unknown'. Left
            # as scraped, the artwork lookup searches Deezer for "Unknown".
            'artist': name if is_artist_chart else credit,
            'image_url': _billboard_art(art),
            'image_url_lg': _billboard_art(art, '344x344'),
            'tags': [],
        }

    chart_songs = [_row(r) for _, r in rows.iterrows()]

    return render_template(
        'chart.html',
        mode='yearend',
        yearend_years=years,
        selected_year=selected_year,
        chart_songs=chart_songs,
        available_dates=[],
        selected_date=None,
        chart=CHARTS.get(chart_key, {}),
        chart_key=chart_key,
    )


def _song_chart_page(source_df, available_dates, endpoint, template):
    """Shared weekly song-chart page renderer (Hot 100 and the global charts)."""
    # Every chart route funnels through here, so the year-end view is picked up
    # for all of them from this one branch: no new routes, no registry change.
    if request.args.get('view') == 'yearend':
        return _yearend_chart_page(endpoint)

    # Get the selected date from query params (default to latest)
    selected_date = request.args.get('date', None)

    data = source_df.copy()
    data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
    data = data.dropna(subset=['Date'])

    # If no date selected, use the latest
    if not selected_date and available_dates:
        selected_date = available_dates[0]

    # Get chart data for selected date
    chart_songs = []
    if selected_date:
        selected_date_dt = pd.to_datetime(selected_date, errors='coerce')
        if pd.isna(selected_date_dt):
            flash('Invalid date', 'error')
            return redirect(url_for(endpoint))
        selected_date = selected_date_dt.strftime('%Y-%m-%d')
        date_data = data[data['Date'] == selected_date_dt]
        if date_data.empty and available_dates:
            avail = pd.to_datetime(pd.Series(available_dates))
            selected_date_dt = avail.iloc[(avail - selected_date_dt).abs().argmin()]
            selected_date = selected_date_dt.strftime('%Y-%m-%d')
            date_data = data[data['Date'] == selected_date_dt]
        date_data = date_data.sort_values('Rank')
        date_data = date_data.dropna(subset=['Rank'])

        # PRE-CALCULATE cumulative weeks for all songs on this chart (PERFORMANCE OPTIMIZATION)
        # Filter data up to selected date once
        historical_data = data[data['Date'] <= selected_date_dt].copy()
        historical_data['Song_Clean'] = historical_data['Song'].str.strip().str.casefold()
        historical_data['Artist_Clean'] = _primary_artist_col(historical_data['Artist'])

        # Restrict history to the song/artist pairs actually on this chart, so the
        # lookups below materialize ~100 keys instead of one per pair that ever charted.
        chart_pairs = set(zip(date_data['Song'].str.strip().str.casefold(),
                              _primary_artist_col(date_data['Artist'])))
        pair_index = pd.MultiIndex.from_frame(historical_data[['Song_Clean', 'Artist_Clean']])
        historical_data = historical_data[pair_index.isin(chart_pairs)].copy()

        # Chart dates per exact credit, and per primary-artist key (credit variants merged)
        historical_data['Artist_Full'] = historical_data['Artist'].str.strip().str.casefold()
        exact_dates = historical_data.groupby(['Song_Clean', 'Artist_Full'])['Date'].agg(set).to_dict()
        pair_dates = historical_data.groupby(['Song_Clean', 'Artist_Clean'])['Date'].agg(set).to_dict()
        chart_week_index = {pd.Timestamp(d): i for i, d in enumerate(available_dates)}

        # Rank each week for every (song, primary-artist) so last-week/peak/change are
        # computed from actual chart history, NOT the stored Last Week / Peak Position
        # columns — those are corrupt in pre-2025 data (debuts stored Last Week == Rank
        # and Peak Position == 1, wrongly showing "—" and a #1 peak).
        historical_data['Rank_Num'] = pd.to_numeric(historical_data['Rank'], errors='coerce')
        pair_ranks = {}
        _rank_grp = historical_data.dropna(subset=['Rank_Num']).groupby(
            ['Song_Clean', 'Artist_Clean', 'Date'])['Rank_Num'].min()
        for (s, a, dt), rk in _rank_grp.items():
            pair_ranks.setdefault((s, a), {})[pd.Timestamp(dt)] = int(rk)

        idx_sel = chart_week_index.get(selected_date_dt)
        prev_date = pd.Timestamp(available_dates[idx_sel + 1]) if (
            idx_sel is not None and idx_sel + 1 < len(available_dates)) else None

        for _, row in date_data.iterrows():
            # Get song info
            song_name = row['Song'].strip() if pd.notna(row['Song']) else ''
            artist_name = row['Artist'].strip() if pd.notna(row['Artist']) else ''

            # Weeks = all weeks under the exact credit, plus credit-variant weeks only
            # within the contiguous run ending at this chart week — so mid-run credit
            # drift still merges, but a later same-titled version with a different
            # credit doesn't inherit an older version's weeks
            merged_dates = pair_dates.get((song_name.casefold(), primary_artist(artist_name)), set())
            song_dates = set(exact_dates.get((song_name.casefold(), artist_name.casefold()), set()))
            cur = selected_date_dt
            while cur in merged_dates:
                song_dates.add(cur)
                idx = chart_week_index.get(cur)
                if idx is None or idx + 1 >= len(available_dates):
                    break
                cur = pd.Timestamp(available_dates[idx + 1])
            cumulative_weeks = len(song_dates) or 1

            rank = safe_int(row['Rank'], 0)
            ranks_by_date = pair_ranks.get((song_name.casefold(), primary_artist(artist_name)), {})
            # Peak = best (lowest) rank across every week the song has charted up to now
            all_ranks = [rk for d, rk in ranks_by_date.items() if d <= selected_date_dt]
            peak = min(all_ranks) if all_ranks else rank
            # Last week = rank on the immediately preceding chart week, or None if it
            # wasn't on that week (a debut or a re-entry after a gap)
            last_week = ranks_by_date.get(prev_date) if prev_date is not None else None

            song_info = {
                'rank': rank,
                'song': song_name,
                'artist': artist_name,
                'last_week': last_week,
                'peak': peak,
                'weeks': cumulative_weeks,
            }

            # Calculate position change
            if last_week is None:
                # Re-entry if the song ever charted before this week, else a true debut
                charted_before = any(d < selected_date_dt for d in merged_dates)
                song_info['change'] = 're-entry' if charted_before else 'new'
                song_info['change_amount'] = 0
            elif rank < last_week:
                song_info['change'] = 'up'
                song_info['change_amount'] = last_week - rank
            elif rank > last_week:
                song_info['change'] = 'down'
                song_info['change_amount'] = rank - last_week
            else:
                song_info['change'] = 'same'
                song_info['change_amount'] = 0

            # Row state, derived from the history already computed above.
            # new_peak drives the .new-peak row styling; tags render as
            # data-tags, which is what the route tests read to check this
            # derivation against the raw CSV.
            # A debut is NOT a new peak: rank == peak holds trivially on a first
            # week, which would tag most of the lower chart and say nothing. A
            # new peak means the song has charted before and just beat its own
            # best — hence the strict comparison against prior weeks only.
            prior_ranks = [rk for d, rk in ranks_by_date.items() if d < selected_date_dt]
            song_info['new_peak'] = bool(prior_ranks) and rank < min(prior_ranks)

            tags = []
            if song_info['change'] in ('new', 're-entry'):
                tags.append(song_info['change'])
            if song_info['new_peak']:
                tags.append('peak')
            song_info['tags'] = tags

            chart_songs.append(song_info)

    return render_template(
        template,
        available_dates=available_dates,
        selected_date=selected_date,
        chart_songs=chart_songs,
        # Registry entry for the page being rendered. chart.html reads these for
        # its title, heading, and history-API chart param; the older per-chart
        # templates ignore them.
        chart=CHARTS.get(endpoint, {}),
        chart_key=endpoint,
        mode='weekly',
        # Drives the Weekly/Year-End toggle. Empty for a chart with no
        # year-end edition, which is how the toggle stays hidden there.
        yearend_years=YEAREND_YEARS.get(endpoint, []),
    )

@app.route('/top100')
@limiter.exempt
def top100():
    """Hot 100 Weekly Chart Viewer"""
    return _song_chart_page(BILLBOARD_DATA, HOT100_AVAILABLE_DATES, 'top100', 'chart.html')

@app.route('/global200')
@limiter.exempt
def global200():
    """Billboard Global 200 Weekly Chart Viewer"""
    if GLOBAL_200_DATA is None:
        flash('Global 200 data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(GLOBAL_200_DATA, GLOBAL200_AVAILABLE_DATES, 'global200', 'chart.html')

@app.route('/globalexus')
@limiter.exempt
def globalexus():
    """Billboard Global Excl. US Weekly Chart Viewer"""
    if GLOBAL_XUS_DATA is None:
        flash('Global Excl. US data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(GLOBAL_XUS_DATA, GLOBALEXUS_AVAILABLE_DATES, 'globalexus', 'chart.html')

@app.route('/radio')
@limiter.exempt
def radio():
    """Radio Songs Weekly Chart Viewer"""
    if RADIO_DATA is None:
        flash('Radio Songs data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(RADIO_DATA, RADIO_AVAILABLE_DATES, 'radio', 'chart.html')

@app.route('/digital')
@limiter.exempt
def digital():
    """Digital Song Sales Weekly Chart Viewer"""
    if DIGITAL_DATA is None:
        flash('Digital Song Sales data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(DIGITAL_DATA, DIGITAL_AVAILABLE_DATES, 'digital', 'chart.html')

@app.route('/streaming')
@limiter.exempt
def streaming():
    """Streaming Songs Weekly Chart Viewer"""
    if STREAMING_DATA is None:
        flash('Streaming Songs data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(STREAMING_DATA, STREAMING_AVAILABLE_DATES, 'streaming', 'chart.html')

@app.route('/artist100')
@limiter.exempt
def artist100():
    """Artist 100 Weekly Chart Viewer"""
    if ARTIST100_DATA is None:
        flash('Artist 100 data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(ARTIST100_DATA, ARTIST100_AVAILABLE_DATES, 'artist100', 'chart.html')

@app.route('/pop_airplay')
@limiter.exempt
def pop_airplay():
    """Billboard Pop Airplay Weekly Chart Viewer"""
    if POP_AIRPLAY_DATA is None:
        flash('Pop Airplay data is not available', 'error')
        return redirect(url_for('top100'))
    return _song_chart_page(POP_AIRPLAY_DATA, POP_AIRPLAY_AVAILABLE_DATES, 'pop_airplay', 'chart.html')

# ── Genre/format airplay charts ─────────────────────────────────────────────
# Registered from the registry against the shared chart.html. The `key=key`
# default argument is required: without it Python's late closure binding would
# make every one of these routes serve whichever chart the loop ended on — a bug
# that renders perfectly and is easy to miss.
for _key in ('adult_pop', 'adult_contemporary', 'rhythmic', 'country_airplay', 'alternative',
             'rnb_hiphop', 'dance_airplay', 'adult_rnb', 'mainstream_rock', 'adult_alternative',
             'rock_airplay', 'christian_airplay', 'gospel_airplay', 'latin_pop_airplay',
             'heatseekers', 'dance_sales', 'bubbling',
             'canadian_hot100', 'dance_electronic', 'japan_hot100', 'uk_songs',
             'rock_songs', 'gospel', 'christian', 'country_songs', 'rnb_songs', 'latin_songs',
             # kind='album' renders correctly through chart.html, which only
             # special-cases 'artist'; the album title takes the song slot.
             'top_album_sales'):
    def _format_chart_page(key=_key):
        df, dates = CHART_DATA[key]
        if df is None:
            flash(f"{CHARTS[key]['label']} data is not available", 'error')
            return redirect(url_for('top100'))
        return _song_chart_page(df, dates, key, 'chart.html')

    app.add_url_rule(f'/{_key}', _key, limiter.exempt(_format_chart_page))
del _key

@app.route('/albums200')
@limiter.exempt
def albums200():
    """Billboard 200 Weekly Albums Chart Viewer"""
    if BILLBOARD_200_DATA is None:
        flash('Billboard 200 data is not available', 'error')
        return redirect(url_for('search'))

    # Was a near-copy of _song_chart_page that stopped just short of the
    # 2026-07-29 debut/peak fix, so it still read the stored Last Week / Peak
    # Position columns — corrupt in pre-2025 rows, where a debut was written as
    # Last Week == Rank and Peak Position == 1. Sharing the renderer fixes those
    # values on this chart too and gives it the row filters for free.
    return _song_chart_page(BILLBOARD_200_DATA, ALBUMS200_AVAILABLE_DATES,
                            'albums200', 'billboard200.html')

# Bubbling Under is the Hot 100's feeder chart, so each side answers a question
# about the other: did this entry graduate, and did that hit start down there?
CROSSOVER_CHART = {'bubbling': 'top100', 'top100': 'bubbling'}


def _crossover_run(chart_key, song, artist, debut):
    """This song's run on the paired chart, or None if it never charted there.

    The join casefolds BOTH fields. Scraped artist casing drifts week to week
    ('The Kid LAROI' vs 'The Kid Laroi') — the 2026-07-01 bug — and an exact-key
    join silently returns nothing, which looks identical to a song that genuinely
    never crossed over.
    """
    other = CROSSOVER_CHART.get(chart_key)
    df, _ = CHART_DATA.get(other, (None, None))
    if df is None or not len(df) or other not in CHART_DT:
        return None

    # Filter on the title first: matching titles are a handful of rows, so the
    # credit normalization runs over those instead of the whole chart.
    same_title = df[df['Song'].astype(str).str.strip().str.casefold() == song.strip().casefold()]
    if same_title.empty:
        return None
    match = same_title[
        same_title['Artist'].map(primary_artist, na_action='ignore') == primary_artist(artist)
    ]
    if match.empty:
        return None

    when = CHART_DT[other].loc[match.index].dropna()
    ranks = pd.to_numeric(match['Rank'], errors='coerce').dropna()
    if when.empty or ranks.empty:
        return None

    first = when.min()
    return {
        'chart': other,
        'label': CHARTS[other]['label'],
        'debut': first.strftime('%Y-%m-%d'),
        'peak': int(ranks.min()),
        'weeks': int(when.nunique()),
        # Whether that run began after this one — the difference between "reached
        # the Hot 100" and "was already there".
        'later': bool(debut is not None and pd.notna(debut) and first > debut),
    }


@app.route('/api/song-history')
@limiter.exempt
def get_song_history():
    """Get full chart history for a specific song (using query parameters to support slashes in names)"""
    artist = request.args.get('artist', '')
    song = request.args.get('song', '')

    if not artist or not song:
        return jsonify({'error': 'Missing artist or song parameter'}), 400

    chart = request.args.get('chart', '')
    # Registry-driven so a new chart cannot silently fall through to the Hot 100
    # and show the wrong song's history. 'popairplay' is the pre-registry param
    # name still emitted by templates that have not moved to chart.html.
    chart = {'popairplay': 'pop_airplay'}.get(chart, chart)
    source, _ = CHART_DATA.get(chart, (None, None))
    if source is None:
        # Name the fallback rather than leaving `chart` empty: everything below
        # is then reading the Hot 100 under the key that says so.
        chart = 'top100'
        source = BILLBOARD_DATA
    data = source.copy()
    data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
    data = data.dropna(subset=['Date'])

    # Clean and filter. Artist matches on primary_artist so a mid-run credit change
    # ("Tame Impala" -> "Tame Impala & JENNIE") still returns the full chart run.
    data['Song_Clean'] = data['Song'].str.strip().str.lower()
    # na_action='ignore' is load-bearing, not tidiness: Artist is a category
    # column, and pandas' Categorical.map calls the mapper with np.nan whenever
    # the column has NAs. Adult Contemporary has 20 blank-artist rows, so
    # without this a primary_artist that assumes a string 500s every song
    # history request for the whole chart.
    data['Artist_Clean'] = data['Artist'].map(primary_artist, na_action='ignore')

    song_data = data[
        (data['Song_Clean'] == song.strip().lower()) &
        (data['Artist_Clean'] == primary_artist(artist))
    ].copy()

    if song_data.empty:
        return jsonify({'error': 'No history found'}), 404

    # Sort by date
    song_data = song_data.sort_values('Date')

    valid = song_data.dropna(subset=['Rank'])
    history = [
        {'date': d.strftime('%Y-%m-%d'), 'rank': int(r), 'weeks': i}
        for i, (d, r) in enumerate(zip(valid['Date'], valid['Rank']), start=1)
    ]

    # Get stats
    rank_clean = song_data['Rank'].dropna()
    peak = int(rank_clean.min()) if not rank_clean.empty else 100
    total_weeks = len(song_data)

    return jsonify({
        'history': history,
        'peak': peak,
        'total_weeks': total_weeks,
        'song': song,
        'artist': artist,
        'crossover': _crossover_run(chart, song, artist, song_data['Date'].min()),
    })

@app.route('/api/album-history')
@limiter.exempt
def get_album_history():
    """Get full chart history for a specific album (using query parameters to support slashes in names)"""
    if BILLBOARD_200_DATA is None:
        return jsonify({'error': 'Billboard 200 data not available'}), 404

    artist = request.args.get('artist', '')
    album = request.args.get('album', '')

    if not artist or not album:
        return jsonify({'error': 'Missing artist or album parameter'}), 400

    data = BILLBOARD_200_DATA.copy()
    data['Date'] = pd.to_datetime(data['Date'], errors='coerce')
    data = data.dropna(subset=['Date'])

    # Clean and filter (Song column contains album names in Billboard 200 data)
    data['Album_Clean'] = data['Song'].str.strip().str.lower()
    data['Artist_Clean'] = data['Artist'].str.strip().str.lower()

    album_data = data[
        (data['Album_Clean'] == album.lower()) &
        (data['Artist_Clean'] == artist.lower())
    ].copy()

    if album_data.empty:
        return jsonify({'error': 'No history found'}), 404

    # Sort by date
    album_data = album_data.sort_values('Date')

    valid = album_data.dropna(subset=['Rank'])
    history = [
        {'date': d.strftime('%Y-%m-%d'), 'rank': int(r), 'weeks': i}
        for i, (d, r) in enumerate(zip(valid['Date'], valid['Rank']), start=1)
    ]

    # Get stats
    rank_clean = album_data['Rank'].dropna()
    peak = int(rank_clean.min()) if not rank_clean.empty else 200
    total_weeks = len(album_data)

    return jsonify({
        'history': history,
        'peak': peak,
        'total_weeks': total_weeks,
        'album': album,
        'artist': artist
    })

@app.route('/api/hot100/current')
@limiter.exempt
def get_current_number_one():
    """API endpoint for Wio Terminal - returns current #1 song"""
    try:
        # Get the most recent date
        latest_date = BILLBOARD_DATA['Date'].max()

        # Get #1 song for that date
        current_number_one = BILLBOARD_DATA[
            (BILLBOARD_DATA['Date'] == latest_date) &
            (BILLBOARD_DATA['Rank'] == 1)
        ]

        if current_number_one.empty:
            return jsonify({'error': 'No #1 song found'}), 404

        song_data = current_number_one.iloc[0]

        return jsonify({
            'song': song_data['Song'],
            'artist': song_data['Artist'],
            'weeks_at_1': int(BILLBOARD_DATA[
                (BILLBOARD_DATA['Song'].astype(str).str.strip() == str(song_data['Song']).strip()) &
                (BILLBOARD_DATA['Artist'].astype(str).str.strip() == str(song_data['Artist']).strip()) &
                (BILLBOARD_DATA['Rank'] == 1)
            ].shape[0]) or 1,
            'image_url': song_data['Image URL'] if 'Image URL' in song_data.index and pd.notna(song_data['Image URL']) else '',
            'date': str(latest_date)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

def _song_pivot(df):
    """One artist's rows as songs-across-the-top, weeks-down-the-side.

    Shared by both CSV exports so a song is columned the same way whether you
    downloaded one artist or a comparison. Returns (pairs, headers, pivot,
    images): `pairs` are the (song, primary artist) column keys in debut
    order, `headers` their display titles, `pivot` a Date-indexed frame of
    ranks, `images` a pair -> URL map. `df` must already be filtered to the
    artist, with Date parsed.
    """
    df = df.dropna(subset=['Date']).copy()
    df['Song'] = df['Song'].astype(str)
    df['Rank'] = pd.to_numeric(df['Rank'], errors='coerce')

    df['Artist'] = df['Artist'].astype(str)
    # Key columns by (song, primary artist) so same-titled songs by different
    # artists stay separate, but a mid-run credit change (remix adds a feature,
    # e.g. "Tame Impala" -> "Tame Impala & JENNIE") stays ONE column.
    df['PrimaryArtist'] = df['Artist'].map(primary_artist)
    pairs = df.groupby(['Song', 'PrimaryArtist'], observed=True)['Date'].min().sort_values().index.tolist()
    pivot = df.pivot_table(index='Date', columns=['Song', 'PrimaryArtist'], values='Rank',
                           aggfunc='min', observed=True).sort_index()

    # For disambiguation, show the latest (fullest) credit for each pair
    latest_credit = df.sort_values('Date').groupby(['Song', 'PrimaryArtist'], observed=True)['Artist'].last()
    title_counts = {}
    for s, a in pairs:
        title_counts[s] = title_counts.get(s, 0) + 1
    headers = [s if title_counts[s] == 1 else f"{s} ({latest_credit.get((s, a), a)})" for s, a in pairs]

    images = {}
    if 'Image URL' in df.columns:
        for pair, grp in df.groupby(['Song', 'PrimaryArtist'], observed=True):
            urls = grp['Image URL'].dropna()
            urls = urls[urls.astype(str).str.startswith('http')]
            images[pair] = str(urls.iloc[0]) if len(urls) else ''

    _fill_missing_images(pairs, images, latest_credit)
    return pairs, headers, pivot, images


# Nine charts (every genre/format airplay CSV, which the newer backfill pipeline
# writes) have no Image URL column, so their exports used to emit a blank Image
# row. Fill from two sources, cheapest first: the charts that DO carry image
# URLs, then Deezer — the same lookup the song pages use, so a CSV agrees with
# what the site shows.
_local_image_index = None


def _build_local_image_index():
    """(song, primary artist) -> artwork URL, from every in-memory chart frame
    that carries an Image URL column. Built once, lazily: it costs a pass over
    ~500k rows and only the CSV exports need it."""
    index = {}
    for key, (frame, _dates) in CHART_DATA.items():
        if frame is None or 'Image URL' not in frame.columns:
            continue
        sub = frame[['Song', 'Artist', 'Image URL']].dropna(subset=['Image URL'])
        sub = sub[sub['Image URL'].astype(str).str.startswith('http')]
        for song, artist, url in sub.itertuples(index=False):
            index.setdefault((str(song).casefold().strip(), primary_artist(str(artist))), str(url))
    return index


# Live lookups are capped per export so one download of a prolific artist cannot
# turn into hundreds of serial API calls. Cache hits do not count against it, so
# a repeated export keeps filling in past the cap.
_DEEZER_EXPORT_LOOKUPS = 120


def _fill_missing_images(pairs, images, latest_credit):
    """Fill blank entries of `images` in place. Returns nothing; blanks that
    survive both sources stay blank rather than guessing."""
    missing = [p for p in pairs if not images.get(p)]
    if not missing:
        return

    global _local_image_index
    if _local_image_index is None:
        _local_image_index = _build_local_image_index()

    still_missing = []
    for pair in missing:
        song, artist = pair
        url = _local_image_index.get((str(song).casefold().strip(), artist))
        if url:
            images[pair] = url
        else:
            still_missing.append(pair)

    budget = _DEEZER_EXPORT_LOOKUPS
    for pair in still_missing:
        song, artist = pair
        # Search on the fullest credit seen for the song, which is what the
        # song pages send and therefore what is already in the cache.
        credit = str(latest_credit.get(pair, artist))
        cache_key = ('track', credit.casefold(), str(song).casefold())
        if _deezer_cache_get(cache_key) is None:
            if budget <= 0:
                continue
            budget -= 1
        try:
            payload = _deezer_track_payload(credit, str(song))
        except Exception as e:      # network/parse — a CSV should still download
            print(f"Deezer image lookup failed for '{song}' by {credit}: {e}")
            continue
        if payload and payload.get('image_url'):
            images[pair] = payload['image_url']


# An export writes only the weeks something was charting, in date order, so a
# song that fell off and came back reads as two adjacent rows that can be years
# apart — nothing on the sheet says a break happened. One filler row per gap,
# date cell a bare dot and every rank blank, sits where the missing weeks would
# have gone. One dot per gap regardless of its length, not one per missing week.
_GAP_MARKER = '.'


def _weeks_with_gaps(dates):
    """Yield each date in order, preceded by None wherever the step from the
    previous date is longer than one week. None means: write a marker row."""
    prev = None
    for date in dates:
        if prev is not None and (date - prev).days > 7:
            yield None
        yield date
        prev = date


def _artist_history_csv(df):
    """One artist's chart history as CSV: songs across the top, an image row,
    then one row per chart week with each song's rank (blank when it was not
    on the chart)."""
    import io, csv as csvmod
    pairs, headers, pivot, images = _song_pivot(df)

    buf = io.StringIO()
    writer = csvmod.writer(buf)
    writer.writerow(['Song'] + headers)
    writer.writerow(['Image'] + [images.get(p, '') for p in pairs])
    for date in _weeks_with_gaps(pivot.index):
        if date is None:
            writer.writerow([_GAP_MARKER] + [''] * len(pairs))
            continue
        row = pivot.loc[date]
        date_str = f"{date.month}/{date.day}/{date.year}"
        writer.writerow([date_str] + [
            (int(row[p]) if pd.notna(row.get(p)) else '') for p in pairs
        ])
    return buf.getvalue()


@app.route('/download-csv/<path:artist_name>')
def download_csv(artist_name):
    """One artist's chart history as CSV.

    Defaults to the Hot 100, which is what the artist report page links to and
    all this route used to serve. ?chart= scopes it to any chart in the
    registry, which is what the versus page's per-artist links use — comparing
    artists on Country Airplay and then being handed their Hot 100 runs would
    not be the same data the page is showing.
    """
    try:
        if not artist_name.strip():
            return jsonify({'error': 'Missing artist name'}), 400
        chart_key = request.args.get('chart', 'top100')
        if chart_key not in CHARTS:
            return jsonify({'error': 'Unknown chart'}), 400
        df, _dates = CHART_DATA.get(chart_key, (None, None))
        if df is None or not len(df):
            return jsonify({'error': 'Chart data is not available'}), 400

        rows = df[artist_match_mask(df['Artist'], artist_name)].copy()
        if rows.empty:
            return jsonify({'error': 'No data for artist'}), 404
        # Dates were parsed once at startup; re-parsing the subset here cost
        # more than building the pivot does.
        rows['Date'] = CHART_DT[chart_key].loc[rows.index]

        safe_name = re.sub(r'[^\w\s-]', '', artist_name).strip().replace(' ', '_')
        # The Hot 100 filename is left as it was, so existing links and any
        # saved files keep their names.
        suffix = '' if chart_key == 'top100' else f'_{chart_key}'
        return Response(
            _artist_history_csv(rows),
            mimetype='text/csv',
            headers={'Content-Disposition':
                     f'attachment; filename="{safe_name}{suffix}_Chart_History.csv"'}
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/<path:artist_name>/<report_type>')
def download_excel(artist_name, report_type='songs'):
    """Download Excel file for artist - songs or albums"""
    try:
        if report_type not in ('songs', 'albums'):
            return jsonify({'error': 'Invalid report type'}), 400
        if not artist_name.strip():
            return jsonify({'error': 'Missing artist name'}), 400
        safe_name = re.sub(r'[^\w\s-]', '', artist_name).strip().replace(' ', '_')
        if report_type == 'albums':
            output_file, error = process_album_data_excel(artist_name)
            filename = f'{safe_name}_Albums_Chart_History.xlsx'
        else:
            output_file, error = process_billboard_data(artist_name)
            filename = f'{safe_name}_Songs_Chart_History.xlsx'

        if error:
            return jsonify({'error': error}), 404

        return send_file(
            output_file,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("\n" + "="*60)
    print("U.S. Top Charts Analyzer - Web App")
    print("="*60)
    print("\nStarting server...")
    port = int(os.environ.get('PORT', 5001))
    debug_mode = os.environ.get('FLASK_ENV', 'production') == 'development'
    print(f"Open your browser and go to: http://localhost:{port}")
    print("\nPress CTRL+C to stop the server")
    print("="*60 + "\n")
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
